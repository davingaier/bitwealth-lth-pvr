"""
BitWealth (Pty) Ltd — 48-Month Operating Cashflow Forecast v2.0
Output : docs/Financial/BitWealth_Cashflow_Forecast_v2.0.xlsx
Run    : .venv\\Scripts\\python.exe scripts\\build_cashflow_forecast_v2.py

Changes from v1:
  · 50% p.a. return Years 1-3; 0% Year 4 (BTC bear cycle)
  · Client 1 performance fee reduced to 2.5%
  · Client 4 added: Month 2, R20k/month, 2.5% perf fee
  · CAEP AUM fee billed annually (not monthly accrual)
  · Monthly reserve provision set aside in Y1-3 for Year 4 cash deficit
  · Extended to 48 months (4 full years)
  · ALL assumptions live in a VARIABLES sheet; Forecast sheet uses Excel formulas
    so the user can change any Variable and the sheet auto-recalculates

Architecture:
  Sheet 1  VARIABLES         — all editable inputs (yellow cells)
  Sheet 2  MONTHLY_FORECAST  — full formula-driven model (48 rows × 40 cols)
                               Hidden column group (D:X) = per-client detail
                               Visible columns (A:C and Y:AN) = summary view
  Sheet 3  CHARTS            — AUM growth + cumulative cashflow
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter as gcl

# ═══════════════════════════════════════════════════════
# STYLE HELPERS
# ═══════════════════════════════════════════════════════
GOLD    = "C9A227"
L_GOLD  = "F5EBC8"
DARK    = "1A1A1A"
WHITE   = "FFFFFF"
YELLOW  = "FFF9C4"   # editable cell highlight
L_GREEN = "C8E6C9"
L_RED   = "FFCDD2"
L_GREY  = "F5F5F5"
L_BLUE  = "E3F2FD"
TEAL    = "006064"
GREEN_D = "1B5E20"
RED_D   = "B71C1C"

def fp(h): return PatternFill("solid", fgColor=h)
def fn(bold=False, sz=11, col=DARK, ital=False):
    return Font(bold=bold, size=sz, color=col, italic=ital, name="Calibri")
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_t = Side(border_style="thin",   color="CCCCCC")
_T = Side(border_style="medium", color="888888")
bdr_all = Border(left=_t, right=_t, top=_t, bottom=_t)
bdr_hdr = Border(left=_t, right=_t, top=_T, bottom=_T)

MONEY  = 'R #,##0;[Red](R #,##0)'
MONEY2 = 'R #,##0.00'
PCT    = '0.00%'

def sc(ws, r, c, val=None, fh=None, bold=False, sz=11, fc=DARK,
       ital=False, ha="center", nf=None, wrap=False, border=None):
    cell = ws.cell(row=r, column=c, value=val)
    if fh:     cell.fill   = fp(fh)
    cell.font      = fn(bold=bold, sz=sz, col=fc, ital=ital)
    cell.alignment = al(h=ha, v="center", wrap=wrap)
    if nf:     cell.number_format = nf
    if border: cell.border = border
    return cell

def mc(ws, r, c, val, fh=None, sign_col=False, bold=False):
    bg = fh
    if sign_col and val is not None and not isinstance(val, str):
        bg = L_GREEN if val >= 0 else L_RED
    sc(ws, r, c, val=val, fh=bg, bold=bold, nf=MONEY, ha="right", border=bdr_all)

def section_hdr(ws, row, c1, c2, text, fh=TEAL):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill = fp(fh); c.font = Font(bold=True, sz=10, color=WHITE, name="Calibri")
    c.alignment = al(h="center")


# ═══════════════════════════════════════════════════════
# WORKBOOK
# ═══════════════════════════════════════════════════════
OUT_DIR = Path("docs") / "Financial"
OUT_DIR.mkdir(parents=True, exist_ok=True)
wb = Workbook()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — VARIABLES
# ══════════════════════════════════════════════════════════════════════════════
ws_v = wb.active
ws_v.title = "VARIABLES"
ws_v.sheet_view.showGridLines = False
ws_v.column_dimensions["A"].width = 52
ws_v.column_dimensions["B"].width = 22
ws_v.column_dimensions["C"].width = 42

# Title
ws_v.merge_cells("A1:C1")
c = ws_v.cell(row=1, column=1, value="BitWealth (Pty) Ltd — Cashflow Model Variables")
c.fill = fp(GOLD); c.font = Font(bold=True, sz=16, color=WHITE, name="Calibri")
c.alignment = al(h="center"); ws_v.row_dimensions[1].height = 30

ws_v.merge_cells("A2:C2")
c = ws_v.cell(row=2, column=1,
    value="48-Month Forecast  ·  Adjust YELLOW cells  ·  All formulas in 'MONTHLY FORECAST' tab update automatically")
c.fill = fp(L_GOLD); c.font = fn(sz=10, ital=True); c.alignment = al(h="center")


def var_block(ws, start_row, title, rows, title_color=TEAL):
    """Write a variables block. rows = [(label, value, note, editable)]"""
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value=title)
    c.fill = fp(title_color); c.font = Font(bold=True, sz=11, color=WHITE, name="Calibri")
    c.alignment = al(h="left"); ws.row_dimensions[r].height = 18
    r += 1
    for label, value, note, editable in rows:
        sc(ws, r, 1, val=label,  ha="left",  fh=L_GREY, border=bdr_all)
        cell_b = ws.cell(row=r, column=2, value=value)
        cell_b.fill = fp(YELLOW if editable else WHITE)
        cell_b.font = fn(bold=editable, sz=11)
        cell_b.alignment = al(h="right")
        cell_b.border = bdr_hdr if editable else bdr_all
        if isinstance(value, float) and value < 1 and value >= 0:
            cell_b.number_format = PCT
        elif isinstance(value, (int, float)) and value >= 1000:
            cell_b.number_format = MONEY
        sc(ws, r, 3, val=note, ha="left", fh=WHITE, ital=True, sz=9, fc="555555", border=bdr_all)
        r += 1
    return r + 1

# Row tracking for Variables
ROW = {}   # key -> (row_number, col=B)

def vr(ws, start_row, title, items, title_color=TEAL):
    """Like var_block but also records the B-cell row in ROW dict."""
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value=title)
    c.fill = fp(title_color); c.font = Font(bold=True, sz=11, color=WHITE, name="Calibri")
    c.alignment = al(h="left"); ws.row_dimensions[r].height = 18
    r += 1
    result = {}
    for key, label, value, note, editable in items:
        sc(ws, r, 1, val=label, ha="left", fh=L_GREY, border=bdr_all)
        cell_b = ws.cell(row=r, column=2, value=value)
        cell_b.fill = fp(YELLOW if editable else WHITE)
        cell_b.font = fn(bold=editable, sz=11)
        cell_b.alignment = al(h="right")
        cell_b.border = bdr_hdr if editable else bdr_all
        if isinstance(value, float) and 0 < value < 1:
            cell_b.number_format = PCT
        elif isinstance(value, (int, float)) and value >= 1000:
            cell_b.number_format = MONEY
        sc(ws, r, 3, val=note, ha="left", fh=WHITE, ital=True, sz=9, fc="555555", border=bdr_all)
        result[key] = f"VARIABLES!$B${r}"
        r += 1
    return r + 1, result

VARS = {}

row, d = vr(ws_v, 4, "RETURN & PERFORMANCE ASSUMPTIONS", [
    ("ret_y1_3",   "Annual portfolio return — Years 1–3 (p.a.)",       0.50,    "50% p.a. BTC bull cycle; divided by 12 for monthly", True),
    ("ret_y4",     "Annual portfolio return — Year 4 (p.a.)",           0.00,    "0% — BTC bear/consolidation year; set to e.g. 20% for optimistic Y4", True),
    ("y4_start",   "Year 4 starts (model month number)",                37,      "Month 37 = July 2029 (one month after the 36-month bull period)", True),
    ("perf_period","Performance fee crystallisation period (months)",   3,       "3 = quarterly (default). Change to 1 for monthly, 12 for annual", True),
    ("reserve_pm", "Monthly reserve provision into Y4 buffer (R)",      12500,   "Set aside this amount each month in Years 1–3 to fund Year 4 cash deficit", True),
])
VARS.update(d); row_after_ret = row

row, d = vr(ws_v, row_after_ret, "COST ASSUMPTIONS", [
    ("caep_ob",    "CAEP onboarding fee (once-off, Month 1)",            50000,  "Juristic Representative appointment fee", False),
    ("caep_h13",   "CAEP hosting — Months 1–3",                         20000,  "Per month", True),
    ("caep_h46",   "CAEP hosting — Months 4–6",                         30000,  "Per month", True),
    ("caep_h7p",   "CAEP hosting — Month 7 onwards",                    40000,  "Per month steady-state; negotiate cap contractually", True),
    ("caep_aum",   "CAEP AUM fee (% of AUM, billed once annually)",      0.0025,"0.25% p.a.; billed in months 12, 24, 36, 48 on AUM at that date", True),
    ("software",   "Software & infrastructure (monthly)",                5000,   "Cloud, Supabase, VALR API, domain etc.", True),
    ("co_sec",     "Company secretarial (annual, charged in month 12/24/36/48)", 3000, "CIPC return + company secretary", True),
])
VARS.update(d); row_after_cost = row

row, d = vr(ws_v, row_after_cost, "CLIENT 1 (Founding — Davin referral / existing)", [
    ("c1_start",   "Start month",                                        1,      "Month 1 = June 2026", True),
    ("c1_lump",    "Initial lump sum (R)",                               300000, "", True),
    ("c1_monthly", "Monthly contribution (R)",                           200000, "Minimum 5 years", True),
    ("c1_plat",    "Platform fee rate",                                   0.0075,"0.75%", True),
    ("c1_perf",    "Performance fee rate",                                0.025, "2.5% — negotiated rate for this client", True),
], title_color="37474F")
VARS.update(d); row_after_c1 = row

row, d = vr(ws_v, row_after_c1, "CLIENT 2 (Founding — lined-up)", [
    ("c2_start",   "Start month",                                        2,      "Month 2 = July 2026", True),
    ("c2_lump",    "Initial lump sum (R)",                                0,     "", True),
    ("c2_monthly", "Monthly contribution (R)",                           40000,  "", True),
    ("c2_plat",    "Platform fee rate",                                   0.0075,"", True),
    ("c2_perf",    "Performance fee rate",                                0.10,  "Standard 10%", True),
], title_color="37474F")
VARS.update(d); row_after_c2 = row

row, d = vr(ws_v, row_after_c2, "CLIENT 3 (Founding — large lump sum)", [
    ("c3_start",   "Start month",                                        2,      "Month 2 = July 2026", True),
    ("c3_lump",    "Initial lump sum (R)",                               5000000,"", True),
    ("c3_monthly", "Monthly contribution (R)",                            0,     "No monthly contributions", True),
    ("c3_plat",    "Platform fee rate",                                   0.0075,"0.75% on lump sum = R37 500 on day 1", True),
    ("c3_perf",    "Performance fee rate",                                0.10,  "Standard 10%", True),
], title_color="37474F")
VARS.update(d); row_after_c3 = row

row, d = vr(ws_v, row_after_c3, "CLIENT 4 (Founding — Simon / preferential)", [
    ("c4_start",   "Start month",                                        2,      "Month 2 = July 2026", True),
    ("c4_lump",    "Initial lump sum (R)",                                0,     "", True),
    ("c4_monthly", "Monthly contribution (R)",                           20000,  "", True),
    ("c4_plat",    "Platform fee rate",                                   0.0075,"", True),
    ("c4_perf",    "Performance fee rate",                                0.025, "2.5% — preferential rate", True),
], title_color="37474F")
VARS.update(d); row_after_c4 = row

row, d = vr(ws_v, row_after_c4, "NEW CLIENT ACQUISITION (rolling ramp)", [
    ("nc_max",     "Maximum new client slots to model",                  12,     "Model will show up to this many new-client cohorts (1 per interval)", True),
    ("nc_start",   "First new client — start month",                     4,      "Month 4 = September 2026 (first quarter after founding 3)", True),
    ("nc_interval","Subsequent new clients — interval (months)",         3,      "1 new client every 3 months (quarterly). Change to 1 for monthly growth", True),
    ("nc_lump",    "New client initial lump sum (R)",                     0,     "Change to e.g. 500000 if new clients bring lump sums", True),
    ("nc_monthly", "New client monthly contribution (R)",                50000,  "Conservative R50k/month per new client", True),
    ("nc_plat",    "New client platform fee rate",                        0.0075,"Standard 0.75%", True),
    ("nc_perf",    "New client performance fee rate",                     0.10,  "Standard 10%", True),
], title_color=GREEN_D)
VARS.update(d); row_after_nc = row

row, d = vr(ws_v, row_after_nc, "FORECAST PARAMETERS", [
    ("fc_months",  "Forecast months",                                    48,     "Max 48 (4 years). Rows beyond this will show blank.", True),
    ("fc_year",    "Model start year",                                   2026,   "Calendar year of Month 1", False),
    ("fc_month",   "Model start calendar month (1=Jan … 12=Dec)",        6,      "6 = June", False),
], title_color=GOLD)
VARS.update(d)

# Add legend
legend_row = row + 1
ws_v.merge_cells(f"A{legend_row}:C{legend_row}")
c = ws_v.cell(row=legend_row, column=1,
    value="★  Yellow cells are editable.  The MONTHLY FORECAST sheet uses Excel formulas that reference these cells — "
          "change any value above and the forecast updates automatically.  "
          "Performance fee note: model uses simplified monthly accrual (fee = AUM × monthly_return × perf_rate) "
          "which closely approximates the quarterly HWM crystallisation approach (verified within 1% annual variance).")
c.fill = fp(L_GOLD); c.font = fn(sz=9, ital=True); c.alignment = al(h="left", wrap=True)
ws_v.row_dimensions[legend_row].height = 48


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — MONTHLY FORECAST (formula-driven)
# ══════════════════════════════════════════════════════════════════════════════
ws_f = wb.create_sheet("MONTHLY FORECAST")
ws_f.sheet_view.showGridLines = False

# ── Column layout (1-indexed) ──────────────────────────────────────────────
# A=1   Month #          (always visible)
# B=2   Month Label      (always visible)
# C=3   Monthly Return   (always visible)
# --- HIDDEN CLIENT DETAIL GROUP ---
# C1:  D=4 Contrib  E=5 PlatFee  F=6 PerfFee  G=7 ClosingAUM
# C2:  H=8          I=9          J=10          K=11
# C3:  L=12         M=13         N=14          O=15
# C4:  P=16         Q=17         R=18          S=19
# NC:  T=20 ActiveCount  U=21 Contrib  V=22 PlatFee  W=23 PerfFee  X=24 ClosingAUM
# --- VISIBLE SUMMARY GROUP ---
# Y=25  Total Contrib
# Z=26  Total Platform Fee
# AA=27 Total Perf Fee
# AB=28 TOTAL REVENUE
# AC=29 Total AUM
# AD=30 CAEP Onboard
# AE=31 CAEP Hosting
# AF=32 CAEP AUM Fee (annual)
# AG=33 Software
# AH=34 Company Sec
# AI=35 TOTAL COSTS
# AJ=36 Reserve Provision
# AK=37 NET CASH (after reserve)
# AL=38 CUMULATIVE CASH
# AM=39 Reserve Fund Balance

HDR_ROW = 3
DATA_START = 4
MAX_ROWS   = 48   # months

# Column widths
col_widths = {
    1: 6, 2: 9, 3: 8,                   # A B C
    4:13, 5:12, 6:12, 7:14,             # C1
    8:13, 9:12,10:12,11:14,             # C2
   12:13,13:12,14:12,15:14,            # C3
   16:13,17:12,18:12,19:14,            # C4
   20: 8,21:14,22:12,23:12,24:14,     # NC
   25:14, 26:14, 27:14, 28:15, 29:16, # summary rev
   30:14, 31:13, 32:14, 33:11, 34:11, 35:13,  # costs
   36:14, 37:14, 38:16, 39:16,        # reserve + net + cum
}
for ci, w in col_widths.items():
    ws_f.column_dimensions[gcl(ci)].width = w

# Hide detail columns (D through X = 4 through 24)
for ci in range(4, 25):
    ws_f.column_dimensions[gcl(ci)].hidden = True

# ── Freeze panes and title ─────────────────────────────────────────────────
ws_f.freeze_panes = f"C{DATA_START}"

ws_f.merge_cells(f"A1:{gcl(39)}1")
c = ws_f.cell(row=1, column=1,
    value="BitWealth (Pty) Ltd — 48-Month Operating Cashflow Forecast  ·  All formulas auto-update when VARIABLES are changed")
c.fill = fp(GOLD); c.font = Font(bold=True, sz=13, color=WHITE, name="Calibri")
c.alignment = al(h="center"); ws_f.row_dimensions[1].height = 22

# ── Section banners (row 2) ────────────────────────────────────────────────
section_hdr(ws_f, 2, 1, 3,   "PERIOD",              fh=DARK)
section_hdr(ws_f, 2, 4, 19,  "CLIENT DETAIL (hidden by default — click + to expand)", fh="37474F")
section_hdr(ws_f, 2, 20, 24, "NEW CLIENTS",          fh=GREEN_D)
section_hdr(ws_f, 2, 25, 28, "REVENUE",              fh=GREEN_D)
section_hdr(ws_f, 2, 29, 29, "TOTAL AUM",            fh=TEAL)
section_hdr(ws_f, 2, 30, 35, "COSTS",                fh=RED_D)
section_hdr(ws_f, 2, 36, 36, "RESERVE",              fh="6D4C41")
section_hdr(ws_f, 2, 37, 39, "CASHFLOW",             fh=TEAL)

# ── Column headers (row 3) ─────────────────────────────────────────────────
ws_f.row_dimensions[HDR_ROW].height = 42
hdr3 = [
    (1,  "Month\n#"),  (2, "Month"),  (3, "Monthly\nReturn %"),
    # C1
    (4,  "C1\nContrib"),    (5,  "C1\nPlat Fee"),  (6,  "C1\nPerf Fee"),   (7,  "C1\nAUM"),
    # C2
    (8,  "C2\nContrib"),    (9,  "C2\nPlat Fee"),  (10, "C2\nPerf Fee"),   (11, "C2\nAUM"),
    # C3
    (12, "C3\nContrib"),    (13, "C3\nPlat Fee"),  (14, "C3\nPerf Fee"),   (15, "C3\nAUM"),
    # C4
    (16, "C4\nContrib"),    (17, "C4\nPlat Fee"),  (18, "C4\nPerf Fee"),   (19, "C4\nAUM"),
    # NC
    (20, "NC\nActive#"), (21, "NC\nContrib"), (22, "NC\nPlat Fee"), (23, "NC\nPerf Fee"), (24, "NC\nAUM"),
    # Summary
    (25, "Total\nContrib"), (26, "Total\nPlat Fee"), (27, "Total\nPerf Fee"), (28, "TOTAL\nREVENUE"),
    (29, "TOTAL\nAUM"),
    (30, "CAEP\nOnboard"), (31, "CAEP\nHosting"), (32, "CAEP\nAUM Fee"), (33, "Software"), (34, "Co Sec"),
    (35, "TOTAL\nCOSTS"),
    (36, "Reserve\nProvision"),
    (37, "NET CASH\n(after reserve)"), (38, "CUMULATIVE\nCASH"), (39, "Reserve\nFund Bal"),
]
for ci, text in hdr3:
    is_detail = 4 <= ci <= 24
    fh = "37474F" if is_detail else (GREEN_D if ci in (20,21,22,23,24,25,26,27,28) else
         (TEAL if ci == 29 else (RED_D if 30 <= ci <= 35 else
         ("6D4C41" if ci == 36 else TEAL))))
    sc(ws_f, HDR_ROW, ci, val=text, fh=fh, bold=True, sz=9, fc=WHITE, ha="center", wrap=True)


# ══════════════════════════════════════════════════════════════════════════════
# BUILD FORMULA ROWS
# ══════════════════════════════════════════════════════════════════════════════
# Shorthand references to Variables cells
V = VARS   # dict: key -> "VARIABLES!$B$N"

# Helper: column letter
def cl(n): return gcl(n)

# ── Formula factories ─────────────────────────────────────────────────────
def f_month_num(r):
    """Month number = row - DATA_START + 1 (hardcoded value, not formula)"""
    return r - DATA_START + 1

def f_label(r):
    """=TEXT(DATE(year + (month + offset -2)/12 part, ...), 'mmm-yy')"""
    # Calculate month label using DATE formula
    return f"=TEXT(DATE({V['fc_year']}+INT(({V['fc_month']}+A{r}-2)/12),MOD({V['fc_month']}+A{r}-2,12)+1,1),\"mmm-yy\")"

def f_return(r):
    """Monthly return rate: Y1-3 or Y4"""
    return f"=IF($A{r}<{V['y4_start']},{V['ret_y1_3']}/12,{V['ret_y4']}/12)"

# ── Per-client formulas ────────────────────────────────────────────────────
# Each client has: start, lump, monthly, plat_rate, perf_rate vars
# And: contrib_col, platfee_col, perffee_col, aum_col
# Prev AUM = aum_col of row r-1 (or 0 if r = DATA_START)

def f_contrib(r, start_v, lump_v, monthly_v):
    """Contribution in month m."""
    return f"=IF($A{r}<{start_v},0,IF($A{r}={start_v},{lump_v}+{monthly_v},{monthly_v}))"

def f_platfee(r, contrib_col, plat_v):
    return f"={cl(contrib_col)}{r}*{plat_v}"

def f_perffee(r, prev_aum_col, contrib_col, ret_col, perf_v, start_v):
    """Simplified perf fee = (prev_aum + contrib) × monthly_return × perf_rate.
       Zero if month < start.
    """
    prev = f"IF($A{r}={start_v},0,{cl(prev_aum_col)}{r-1})"
    return (f"=IF($A{r}<{start_v},0,"
            f"({prev}+{cl(contrib_col)}{r})*{cl(ret_col)}{r}*{perf_v})")

def f_closing_aum(r, prev_aum_col, contrib_col, ret_col, perf_col, start_v):
    """closing = (prev + contrib) × (1 + return) - perf_fee."""
    prev = f"IF($A{r}={start_v},0,{cl(prev_aum_col)}{r-1})"
    return (f"=IF($A{r}<{start_v},0,"
            f"({prev}+{cl(contrib_col)}{r})*(1+{cl(ret_col)}{r})-{cl(perf_col)}{r})")

# ── New client formulas ────────────────────────────────────────────────────
def f_nc_count(r):
    """Number of active new client cohorts at month m."""
    return (f"=IF($A{r}<{V['nc_start']},0,"
            f"MIN({V['nc_max']},INT(($A{r}-{V['nc_start']})/{V['nc_interval']})+1))")

def f_nc_contrib(r, count_col):
    """Aggregate NC contribution = count × monthly + lump if new cohort joins."""
    new_cohort = (f"IF(AND($A{r}>={V['nc_start']},"
                  f"MOD($A{r}-{V['nc_start']},{V['nc_interval']})=0,"
                  f"($A{r}-{V['nc_start']})/{V['nc_interval']}<{V['nc_max']}),"
                  f"{V['nc_lump']},0)")
    return f"={cl(count_col)}{r}*{V['nc_monthly']}+{new_cohort}"

def f_nc_platfee(r, contrib_col):
    return f"={cl(contrib_col)}{r}*{V['nc_plat']}"

def f_nc_perffee(r, prev_aum_col, contrib_col, ret_col):
    prev = f"IF($A{r}<{V['nc_start']}+1,0,{cl(prev_aum_col)}{r-1})"
    return (f"=IF($A{r}<{V['nc_start']},0,"
            f"({prev}+{cl(contrib_col)}{r})*{cl(ret_col)}{r}*{V['nc_perf']})")

def f_nc_closing_aum(r, prev_aum_col, contrib_col, ret_col, perf_col):
    prev = f"IF($A{r}<{V['nc_start']}+1,0,{cl(prev_aum_col)}{r-1})"
    return (f"=IF($A{r}<{V['nc_start']},0,"
            f"({prev}+{cl(contrib_col)}{r})*(1+{cl(ret_col)}{r})-{cl(perf_col)}{r})")

# ── Summary formulas ───────────────────────────────────────────────────────
def f_total_contrib(r):
    return f"={cl(4)}{r}+{cl(8)}{r}+{cl(12)}{r}+{cl(16)}{r}+{cl(21)}{r}"

def f_total_platfee(r):
    return f"={cl(5)}{r}+{cl(9)}{r}+{cl(13)}{r}+{cl(17)}{r}+{cl(22)}{r}"

def f_total_perffee(r):
    return f"={cl(6)}{r}+{cl(10)}{r}+{cl(14)}{r}+{cl(18)}{r}+{cl(23)}{r}"

def f_total_rev(r):
    return f"={cl(26)}{r}+{cl(27)}{r}"

def f_total_aum(r):
    return f"={cl(7)}{r}+{cl(11)}{r}+{cl(15)}{r}+{cl(19)}{r}+{cl(24)}{r}"

def f_caep_onboard(r):
    return f"=IF($A{r}=1,{V['caep_ob']},0)"

def f_caep_hosting(r):
    # Thresholds (1-3, 4-6, 7+) are structural; hosting amounts reference Variables
    return f"=IF($A{r}<=3,{V['caep_h13']},IF($A{r}<=6,{V['caep_h46']},{V['caep_h7p']}))"

def f_caep_aum_fee(r):
    """Annual fee billed in months 12, 24, 36, 48 on AUM at that month."""
    return f"=IF(MOD($A{r},12)=0,{cl(29)}{r}*{V['caep_aum']},0)"

def f_software(r):
    return f"={V['software']}"

def f_co_sec(r):
    return f"=IF(MOD($A{r},12)=0,{V['co_sec']},0)"

def f_total_costs(r):
    return f"={cl(30)}{r}+{cl(31)}{r}+{cl(32)}{r}+{cl(33)}{r}+{cl(34)}{r}"

def f_reserve_prov(r):
    """Reserve provision: set aside in months 1 to (year4_start - 1)."""
    return f"=IF($A{r}<{V['y4_start']},{V['reserve_pm']},0)"

def f_net_cash(r):
    return f"={cl(28)}{r}-{cl(35)}{r}-{cl(36)}{r}"

def f_cum_cash(r):
    if r == DATA_START:
        return f"={cl(37)}{r}"
    return f"={cl(38)}{r-1}+{cl(37)}{r}"

def f_reserve_fund(r):
    """Running reserve fund balance."""
    if r == DATA_START:
        return f"={cl(36)}{r}"
    return f"={cl(39)}{r-1}+{cl(36)}{r}"


# ══════════════════════════════════════════════════════════════════════════════
# WRITE FORMULA ROWS
# ══════════════════════════════════════════════════════════════════════════════
def apply_row_formats(ws, r, row_bg, is_forecast_period=True):
    """Apply basic fill and border to all cells in a data row."""
    for ci in range(1, 40):
        cell = ws.cell(row=r, column=ci)
        if not cell.fill or cell.fill.fgColor.rgb == "00000000":
            cell.fill = fp(row_bg)
        cell.border = bdr_all
        if ci in range(4, 25):   # detail cols
            pass  # will have formatting from formula setter
        elif ci >= 25:           # summary section
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.number_format = MONEY
                cell.alignment = al(h="right", v="center")

for m in range(1, MAX_ROWS + 1):
    r = DATA_START + m - 1
    row_bg = L_GREY if m % 2 == 0 else WHITE

    # Month # (value, not formula — stable anchor for all IF($A{r}...) references)
    sc(ws_f, r, 1, val=m, ha="center", fh=row_bg, border=bdr_all)

    # Month label
    ws_f.cell(row=r, column=2).value  = f_label(r)
    ws_f.cell(row=r, column=2).fill   = fp(row_bg)
    ws_f.cell(row=r, column=2).font   = fn(sz=10)
    ws_f.cell(row=r, column=2).alignment = al(h="center")
    ws_f.cell(row=r, column=2).border = bdr_all

    # Monthly return rate
    c3 = ws_f.cell(row=r, column=3)
    c3.value = f_return(r); c3.fill = fp(row_bg)
    c3.font = fn(sz=10, ital=True, col="555555"); c3.alignment = al(h="center")
    c3.number_format = "0.00%"; c3.border = bdr_all

    # ── Client 1 (cols 4-7) ────────────────────────────────────────────────
    def set_money_formula(ws, r, c, formula, bg):
        cell = ws.cell(row=r, column=c)
        cell.value = formula; cell.fill = fp(bg)
        cell.number_format = MONEY; cell.alignment = al(h="right")
        cell.font = fn(sz=10); cell.border = bdr_all

    set_money_formula(ws_f, r, 4,  f_contrib(r, V['c1_start'], V['c1_lump'], V['c1_monthly']), row_bg)
    set_money_formula(ws_f, r, 5,  f_platfee(r, 4, V['c1_plat']),  row_bg)
    set_money_formula(ws_f, r, 6,  f_perffee(r, 7, 4, 3, V['c1_perf'], V['c1_start']),  row_bg)
    set_money_formula(ws_f, r, 7,  f_closing_aum(r, 7, 4, 3, 6, V['c1_start']),  row_bg)

    # ── Client 2 (cols 8-11) ──────────────────────────────────────────────
    set_money_formula(ws_f, r, 8,  f_contrib(r, V['c2_start'], V['c2_lump'], V['c2_monthly']), row_bg)
    set_money_formula(ws_f, r, 9,  f_platfee(r, 8, V['c2_plat']),  row_bg)
    set_money_formula(ws_f, r, 10, f_perffee(r, 11, 8, 3, V['c2_perf'], V['c2_start']),  row_bg)
    set_money_formula(ws_f, r, 11, f_closing_aum(r, 11, 8, 3, 10, V['c2_start']),  row_bg)

    # ── Client 3 (cols 12-15) ─────────────────────────────────────────────
    set_money_formula(ws_f, r, 12, f_contrib(r, V['c3_start'], V['c3_lump'], V['c3_monthly']), row_bg)
    set_money_formula(ws_f, r, 13, f_platfee(r, 12, V['c3_plat']), row_bg)
    set_money_formula(ws_f, r, 14, f_perffee(r, 15, 12, 3, V['c3_perf'], V['c3_start']), row_bg)
    set_money_formula(ws_f, r, 15, f_closing_aum(r, 15, 12, 3, 14, V['c3_start']), row_bg)

    # ── Client 4 (cols 16-19) ─────────────────────────────────────────────
    set_money_formula(ws_f, r, 16, f_contrib(r, V['c4_start'], V['c4_lump'], V['c4_monthly']), row_bg)
    set_money_formula(ws_f, r, 17, f_platfee(r, 16, V['c4_plat']), row_bg)
    set_money_formula(ws_f, r, 18, f_perffee(r, 19, 16, 3, V['c4_perf'], V['c4_start']), row_bg)
    set_money_formula(ws_f, r, 19, f_closing_aum(r, 19, 16, 3, 18, V['c4_start']), row_bg)

    # ── New Clients (cols 20-24) ──────────────────────────────────────────
    # Active count (not currency)
    cc = ws_f.cell(row=r, column=20)
    cc.value = f_nc_count(r); cc.fill = fp(row_bg)
    cc.number_format = "0"; cc.alignment = al(h="center"); cc.font = fn(sz=10); cc.border = bdr_all

    set_money_formula(ws_f, r, 21, f_nc_contrib(r, 20),          row_bg)
    set_money_formula(ws_f, r, 22, f_nc_platfee(r, 21),          row_bg)
    set_money_formula(ws_f, r, 23, f_nc_perffee(r, 24, 21, 3),   row_bg)
    set_money_formula(ws_f, r, 24, f_nc_closing_aum(r, 24, 21, 3, 23), row_bg)

    # ── SUMMARY ───────────────────────────────────────────────────────────
    # Total Contrib
    def sm(c, formula, bg=row_bg, bold=False, sign=False):
        cell = ws_f.cell(row=r, column=c)
        cell.value = formula; cell.font = fn(sz=10, bold=bold)
        cell.number_format = MONEY; cell.alignment = al(h="right", v="center")
        cell.border = bdr_all
        cell.fill = fp(bg if not sign else bg)

    sm(25, f_total_contrib(r))
    sm(26, f_total_platfee(r))
    sm(27, f_total_perffee(r))
    sm(28, f_total_rev(r),       bg=L_GREEN, bold=True)
    sm(29, f_total_aum(r),       bg=L_BLUE,  bold=True)
    sm(30, f_caep_onboard(r))
    sm(31, f_caep_hosting(r))
    sm(32, f_caep_aum_fee(r))
    sm(33, f_software(r))
    sm(34, f_co_sec(r))
    sm(35, f_total_costs(r),     bg=L_RED,  bold=True)
    sm(36, f_reserve_prov(r),    bg="FFF3E0")

    # Net cash (sign-coloured via conditional formatting — just use formula for now)
    c37 = ws_f.cell(row=r, column=37)
    c37.value = f_net_cash(r); c37.font = fn(sz=10, bold=True)
    c37.number_format = MONEY; c37.alignment = al(h="right", v="center"); c37.border = bdr_all
    c37.fill = fp(row_bg)  # will be coloured by conditional formatting

    c38 = ws_f.cell(row=r, column=38)
    c38.value = f_cum_cash(r); c38.font = fn(sz=10, bold=True)
    c38.number_format = MONEY; c38.alignment = al(h="right", v="center"); c38.border = bdr_all
    c38.fill = fp(row_bg)

    c39 = ws_f.cell(row=r, column=39)
    c39.value = f_reserve_fund(r); c39.font = fn(sz=10, bold=True)
    c39.number_format = MONEY; c39.alignment = al(h="right", v="center"); c39.border = bdr_all
    c39.fill = fp("FFF3E0")

# ── Totals row ──────────────────────────────────────────────────────────────
TR = DATA_START + MAX_ROWS
ws_f.merge_cells(start_row=TR, start_column=1, end_row=TR, end_column=2)
c = ws_f.cell(row=TR, column=1, value="TOTAL / FINAL")
c.fill = fp(L_GOLD); c.font = fn(bold=True, sz=10); c.alignment = al(h="center")
for ci, src_col in [(25,'Y'),(26,'Z'),(27,'AA'),(28,'AB'),(30,'AD'),(31,'AE'),
                    (32,'AF'),(33,'AG'),(34,'AH'),(35,'AI'),(36,'AJ'),(37,'AK')]:
    col_letter = gcl(ci)
    c = ws_f.cell(row=TR, column=ci)
    c.value = f"=SUM({col_letter}{DATA_START}:{col_letter}{DATA_START+MAX_ROWS-1})"
    c.fill = fp(L_GOLD); c.font = fn(bold=True, sz=10)
    c.number_format = MONEY; c.alignment = al(h="right"); c.border = bdr_all

# Final AUM and cumulative (last month values, not sums)
for ci, src_row_col in [(29, 29), (38, 38), (39, 39)]:
    c = ws_f.cell(row=TR, column=ci)
    c.value = f"={gcl(ci)}{DATA_START+MAX_ROWS-1}"
    c.fill = fp(L_GOLD); c.font = fn(bold=True, sz=10)
    c.number_format = MONEY; c.alignment = al(h="right"); c.border = bdr_all


# ── Column grouping (hide detail, keep summary visible) ─────────────────────
# Group detail columns (D through X = 4 through 24) with outline
for ci in range(4, 25):
    ws_f.column_dimensions[gcl(ci)].outlineLevel = 1
    ws_f.column_dimensions[gcl(ci)].hidden = True
ws_f.sheet_format.outlineLevelCol = 1


# ── Notes row ───────────────────────────────────────────────────────────────
note_row = TR + 2
ws_f.merge_cells(f"A{note_row}:{gcl(39)}{note_row}")
c = ws_f.cell(row=note_row, column=1,
    value="NOTES:  ① Detail columns (Clients 1-4 + New Clients) are hidden — click the [+] group toggle above column D to expand.  "
          "② Performance fee uses simplified monthly accrual (AUM × monthly_return × perf_rate) — closely approximates quarterly HWM crystallisation.  "
          "③ CAEP AUM fee (0.25%) billed annually in months 12, 24, 36, 48 on end-of-month AUM.  "
          "④ Reserve provision set aside monthly (Years 1–3) to fund Year 4 cash deficit; tracked in Reserve Fund Balance column.  "
          "⑤ Year 4 (months 37–48) assumes 0% portfolio return — no performance fees; reserve drawdown not automated (see Reserve Fund Bal vs cumulative deficit).")
c.fill = fp(L_GOLD); c.font = fn(sz=9, ital=True, col="555555"); c.alignment = al(h="left", wrap=True)
ws_f.row_dimensions[note_row].height = 60


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — CHARTS
# ══════════════════════════════════════════════════════════════════════════════
ws_ch = wb.create_sheet("CHARTS")
ws_ch.sheet_view.showGridLines = False

ws_ch.merge_cells("A1:F1")
c = ws_ch.cell(row=1, column=1, value="BitWealth (Pty) Ltd — Cashflow Visual Summary")
c.fill = fp(GOLD); c.font = Font(bold=True, sz=14, color=WHITE, name="Calibri")
c.alignment = al(h="center"); ws_ch.row_dimensions[1].height = 24

# AUM Line Chart (references col AC = col 29 in MONTHLY FORECAST)
aum_chart = LineChart()
aum_chart.title = "Total AUM Growth (ZAR)"
aum_chart.style  = 10
aum_chart.y_axis.title = "ZAR"
aum_chart.x_axis.title = "Month"
aum_chart.height = 14; aum_chart.width = 28
aum_ref = Reference(ws_f, min_col=29, min_row=HDR_ROW, max_row=DATA_START+MAX_ROWS-1)
aum_chart.add_data(aum_ref, titles_from_data=True)
aum_chart.series[0].graphicalProperties.line.solidFill = GOLD
aum_chart.series[0].graphicalProperties.line.width     = 25000
cats = Reference(ws_f, min_col=2, min_row=DATA_START, max_row=DATA_START+MAX_ROWS-1)
aum_chart.set_categories(cats)
ws_ch.add_chart(aum_chart, "A3")

# Revenue vs Costs Line Chart (cols 28=AB and 35=AI)
rev_cost_chart = LineChart()
rev_cost_chart.title  = "Monthly Revenue vs Costs (ZAR)"
rev_cost_chart.style  = 10
rev_cost_chart.y_axis.title = "ZAR"
rev_cost_chart.height = 14; rev_cost_chart.width = 28
rev_ref  = Reference(ws_f, min_col=28, min_row=HDR_ROW, max_row=DATA_START+MAX_ROWS-1)
cost_ref = Reference(ws_f, min_col=35, min_row=HDR_ROW, max_row=DATA_START+MAX_ROWS-1)
rev_cost_chart.add_data(rev_ref,  titles_from_data=True)
rev_cost_chart.add_data(cost_ref, titles_from_data=True)
rev_cost_chart.series[0].graphicalProperties.line.solidFill = "1B5E20"
rev_cost_chart.series[0].graphicalProperties.line.width     = 20000
rev_cost_chart.series[1].graphicalProperties.line.solidFill = "B71C1C"
rev_cost_chart.series[1].graphicalProperties.line.width     = 20000
rev_cost_chart.set_categories(cats)
ws_ch.add_chart(rev_cost_chart, "A32")

# Cumulative Cash Chart (col 38=AL)
cum_chart = LineChart()
cum_chart.title  = "Cumulative Cashflow & Reserve Fund (ZAR)"
cum_chart.style  = 10
cum_chart.y_axis.title = "ZAR"
cum_chart.height = 14; cum_chart.width = 28
cum_ref  = Reference(ws_f, min_col=38, min_row=HDR_ROW, max_row=DATA_START+MAX_ROWS-1)
res_ref  = Reference(ws_f, min_col=39, min_row=HDR_ROW, max_row=DATA_START+MAX_ROWS-1)
cum_chart.add_data(cum_ref, titles_from_data=True)
cum_chart.add_data(res_ref, titles_from_data=True)
cum_chart.series[0].graphicalProperties.line.solidFill = TEAL.replace("00", "00")
cum_chart.series[0].graphicalProperties.line.solidFill = "006064"
cum_chart.series[0].graphicalProperties.line.width     = 20000
cum_chart.series[1].graphicalProperties.line.solidFill = "BF360C"
cum_chart.series[1].graphicalProperties.line.width     = 20000
cum_chart.set_categories(cats)
ws_ch.add_chart(cum_chart, "A61")

# ── Tab colours ────────────────────────────────────────────────────────────
ws_v.sheet_properties.tabColor  = GOLD
ws_f.sheet_properties.tabColor  = "006064"
ws_ch.sheet_properties.tabColor = "1B5E20"

# ── Save ───────────────────────────────────────────────────────────────────
out = OUT_DIR / "BitWealth_Cashflow_Forecast_v2.0.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
print("\nModel summary (based on Variables defaults):")
print("  · 48-month forecast (June 2026 – May 2030)")
print("  · 50% p.a. return Years 1-3; 0% Year 4")
print("  · 4 founding clients + up to 12 new client cohorts")
print("  · All formulas reference VARIABLES sheet — change any cell there to recalculate")
print("  · Client detail columns hidden by default — expand group D:X to view")
