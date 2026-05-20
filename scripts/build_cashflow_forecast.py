"""
BitWealth (Pty) Ltd — 36-Month Operating Cashflow Forecast
Output : docs/Financial/BitWealth_Cashflow_Forecast_v1.0.xlsx
Run    : .\.venv\Scripts\python.exe scripts\build_cashflow_forecast.py

Assumptions baked in (change in ASSUMPTIONS block):
  · Month 1 = June 2026
  · Platform fee  : 0.75 % on every contribution (lump sum + monthly)
  · Perf fee      : 10 % above HWM, crystallised quarterly (default per client)
  · Portfolio return: 20 % p.a. simple → 1.6667 % per month
  · CAEP onboarding : R 50 000 once-off (Month 1)
  · CAEP hosting    : R 20k M1-3 · R 30k M4-6 · R 40k M7+
  · CAEP AUM fee    : 0.25 % p.a. of end-of-month AUM (monthly accrual — cost)
  · Software        : R 5 000 / month
  · Company sec     : R 3 000 p.a., billed in month 12 / 24 / 36
  · Client 1        : R 300k lump + R 200k/month from M1
  · Client 2        : R 40k/month from M2 (no lump)
  · Client 3        : R 5m lump, no monthly, from M2
  · New clients     : 1 × R 50k/month client from M4, every 3 months thereafter
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════
# 1.  ASSUMPTIONS
# ═══════════════════════════════════════════════════════════
MONTHS          = 36
BASE_YEAR       = 2026
BASE_MO         = 6           # June 2026 = Month 1

PLATFORM_FEE    = 0.0075      # 0.75 % on all contributions
PERF_FEE        = 0.10        # 10 % above HWM
MONTHLY_RET     = 0.20 / 12   # 20 % p.a. simple
PERF_PERIOD     = 3           # crystallise every 3 months (quarterly)

AUM_COST_PA     = 0.0025      # 0.25 % p.a. cost paid to CAEP
CAEP_ONBOARD    = 50_000
SOFTWARE        = 5_000
CO_SEC_ANNUAL   = 3_000       # charged in months 12, 24, 36

CAEP_HOSTING_SCHED = {1: 20_000, 2: 20_000, 3: 20_000,
                       4: 30_000, 5: 30_000, 6: 30_000}
CAEP_HOSTING_BASE  = 40_000   # month 7+

CLIENTS_CONFIG = [
    dict(name="Client 1 (Founding)", start=1, lump=300_000,   monthly=200_000),
    dict(name="Client 2 (Founding)", start=2, lump=0,         monthly=40_000),
    dict(name="Client 3 (Founding)", start=2, lump=5_000_000, monthly=0),
]

NEW_CLIENT_MONTHLY = 50_000
NEW_CLIENT_START   = 4        # first new-client month
NEW_CLIENT_INTERVAL= 3        # add 1 new client every 3 months


# ═══════════════════════════════════════════════════════════
# 2.  SIMULATION
# ═══════════════════════════════════════════════════════════
_MO_NAMES = "Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split()

def mo_label(m):
    """Jun-26, Jul-26, …"""
    total = BASE_MO - 1 + (m - 1)
    yr = BASE_YEAR + total // 12
    mo = total % 12 + 1
    return f"{_MO_NAMES[mo-1]}-{str(yr)[2:]}"

def caep_hosting(m):
    return CAEP_HOSTING_SCHED.get(m, CAEP_HOSTING_BASE)


class ClientTracker:
    def __init__(self, name, start, lump, monthly, period=PERF_PERIOD):
        self.name   = name
        self.start  = start
        self.lump   = lump
        self.monthly= monthly
        self.period = period
        self.aum    = 0.0
        self.hwm    = 0.0
        self.since  = 0     # months since last crystallisation

    def step(self, m):
        if m < self.start:
            return dict(contrib=0., opening=0., ret_zar=0.,
                        closing=0., hwm=0., plat=0., perf=0.)
        contrib  = float(self.lump + self.monthly) if m == self.start else float(self.monthly)
        plat     = contrib * PLATFORM_FEE
        opening  = self.aum
        self.aum += contrib
        self.hwm += contrib       # cost-basis HWM: new money doesn't generate immediate profit
        ret      = self.aum * MONTHLY_RET
        self.aum += ret
        self.since += 1
        perf = 0.0
        if self.since >= self.period:
            if self.aum > self.hwm:
                perf      = (self.aum - self.hwm) * PERF_FEE
                self.aum -= perf
                self.hwm  = self.aum   # reset HWM after fee extraction
            self.since = 0
        return dict(contrib=contrib, opening=opening, ret_zar=ret,
                    closing=self.aum, hwm=self.hwm, plat=plat, perf=perf)


def run_sim():
    clients = [ClientTracker(**c) for c in CLIENTS_CONFIG]
    for i, s in enumerate(range(NEW_CLIENT_START, MONTHS + 1, NEW_CLIENT_INTERVAL)):
        clients.append(ClientTracker(
            name=f"New Client {i+1} (from M{s})", start=s, lump=0, monthly=NEW_CLIENT_MONTHLY))

    rows   = []
    cum    = 0.0
    for m in range(1, MONTHS + 1):
        cd           = [c.step(m) for c in clients]
        total_aum    = sum(c.aum for c in clients)
        total_contrib= sum(d['contrib'] for d in cd)
        total_plat   = sum(d['plat']    for d in cd)
        total_perf   = sum(d['perf']    for d in cd)
        total_rev    = total_plat + total_perf

        c_onboard  = CAEP_ONBOARD   if m == 1        else 0.
        c_host     = float(caep_hosting(m))
        c_aum_fee  = total_aum * AUM_COST_PA / 12
        c_sw       = float(SOFTWARE)
        c_sec      = float(CO_SEC_ANNUAL) if m % 12 == 0 else 0.
        total_cost = c_onboard + c_host + c_aum_fee + c_sw + c_sec

        net  = total_rev - total_cost
        cum += net

        rows.append(dict(
            month=m,        label=mo_label(m),
            cd=cd,          names=[c.name for c in clients],
            aum=total_aum,  contrib=total_contrib,
            plat=total_plat, perf=total_perf, rev=total_rev,
            c_onboard=c_onboard, c_host=c_host, c_aum=c_aum_fee,
            c_sw=c_sw,      c_sec=c_sec,
            cost=total_cost, net=net, cum=cum,
        ))
    return rows, clients


rows, all_clients = run_sim()

first_pos_net = next((r['month'] for r in rows if r['net'] > 0), None)
first_pos_cum = next((r['month'] for r in rows if r['cum'] > 0), None)
peak_deficit  = min(r['cum'] for r in rows)
m36_aum       = rows[-1]['aum']

print(f"  AUM Month 36       : R {m36_aum:,.0f}")
print(f"  First +ve net month: Month {first_pos_net}  ({mo_label(first_pos_net) if first_pos_net else 'n/a'})")
print(f"  Cumulative breakeven: Month {first_pos_cum}  ({mo_label(first_pos_cum) if first_pos_cum else 'n/a'})")
print(f"  Peak cash required : R {abs(peak_deficit):,.0f}")


# ═══════════════════════════════════════════════════════════
# 3.  EXCEL HELPERS
# ═══════════════════════════════════════════════════════════
GOLD     = "C9A227"
L_GOLD   = "F5EBC8"
DARK     = "1A1A1A"
WHITE    = "FFFFFF"
L_GREEN  = "C8E6C9"
L_RED    = "FFCDD2"
L_GREY   = "F5F5F5"
L_BLUE   = "DDEEFF"
ORANGE   = "FF6F00"
TEAL     = "006064"

def fp(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def fnt(bold=False, sz=11, col=DARK, ital=False):
    return Font(bold=bold, size=sz, color=col, italic=ital, name="Calibri")

def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_thin  = Side(border_style="thin",   color="CCCCCC")
_thick = Side(border_style="medium", color="888888")
bdr_all  = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thin)
bdr_top  = Border(left=_thin,  right=_thin,  top=_thick, bottom=_thin)
bdr_bot  = Border(left=_thin,  right=_thin,  top=_thin,  bottom=_thick)

MONEY_FMT = 'R #,##0;[Red](R #,##0)'
MONEY2    = 'R #,##0.00'
NUM_FMT   = '#,##0'


def sc(ws, r, c, val=None, fh=None, bold=False, sz=11, fc=DARK,
       ital=False, ha="center", nf=None, wrap=False, border=None):
    cell = ws.cell(row=r, column=c, value=val)
    if fh:     cell.fill   = fp(fh)
    cell.font      = fnt(bold=bold, sz=sz, col=fc, ital=ital)
    cell.alignment = aln(h=ha, v="center", wrap=wrap)
    if nf:     cell.number_format = nf
    if border: cell.border = border
    return cell


def mc(ws, r, c, val, fh=None, bold=False, sign_color=False):
    """Money cell. sign_color=True → green/red based on value sign."""
    bg = fh
    if sign_color and val is not None:
        bg = L_GREEN if val >= 0 else L_RED
    sc(ws, r, c, val=val, fh=bg, bold=bold, nf=MONEY_FMT, ha="right", border=bdr_all)


def section_header(ws, row, col_start, col_end, text, fh=TEAL):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fp(fh)
    c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    c.alignment = aln(h="center")


# ═══════════════════════════════════════════════════════════
# 4.  BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════
OUT_DIR = Path("docs") / "Financial"
OUT_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()

# ───────────────────────────────────────────────────────────
# SHEET 1: Assumptions
# ───────────────────────────────────────────────────────────
ws_a = wb.active
ws_a.title = "Assumptions"
ws_a.sheet_view.showGridLines = False
ws_a.column_dimensions["A"].width = 40
ws_a.column_dimensions["B"].width = 22
ws_a.column_dimensions["C"].width = 35

# Title
ws_a.merge_cells("A1:C1")
c = ws_a.cell(row=1, column=1, value="BitWealth (Pty) Ltd — Cashflow Forecast: Key Assumptions")
c.fill = fp(GOLD); c.font = Font(bold=True, sz=16, color=WHITE, name="Calibri")
c.alignment = aln(h="center")
ws_a.row_dimensions[1].height = 30

ws_a.merge_cells("A2:C2")
c = ws_a.cell(row=2, column=1,
    value="36-Month Forecast · Baseline scenario · June 2026 – May 2029 · All amounts in ZAR")
c.fill = fp(L_GOLD); c.font = fnt(sz=10, ital=True)
c.alignment = aln(h="center")

def asm_block(ws, start_row, title, params, title_color=TEAL):
    """Write an assumptions block. params = list of (label, value, note)."""
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value=title)
    c.fill = fp(title_color); c.font = Font(bold=True, sz=11, color=WHITE, name="Calibri")
    c.alignment = aln(h="left")
    ws.row_dimensions[r].height = 18
    r += 1
    for label, val, note in params:
        sc(ws, r, 1, val=label, ha="left",  fh=L_GREY, bold=False, border=bdr_all)
        sc(ws, r, 2, val=val,   ha="right", fh=WHITE,  bold=True,  border=bdr_all)
        sc(ws, r, 3, val=note,  ha="left",  fh=WHITE,  ital=True, sz=9, fc="555555", border=bdr_all)
        r += 1
    return r + 1  # blank row gap

row = 4
row = asm_block(ws_a, row, "REVENUE ASSUMPTIONS", [
    ("Platform fee rate",                    "0.75%",          "Charged on ALL contributions (lump sum + monthly)"),
    ("Performance fee rate",                 "10%",            "10% of profits above High Water Mark"),
    ("Performance fee crystallisation",      "Quarterly",      "Every 3 months per client (client may opt for monthly or annual)"),
    ("HWM method",                           "Cost-basis",     "New contributions added to HWM so new money is never immediately profitable"),
    ("Annual portfolio return (assumed)",    "20% p.a.",       "1.6667% per month simple; used to project AUM and performance fee income"),
])

row = asm_block(ws_a, row, "COST ASSUMPTIONS", [
    ("CAEP onboarding fee (once-off)",       "R 50 000",       "Paid Month 1 when Juristic Representative appointment confirmed"),
    ("CAEP hosting — Months 1-3",            "R 20 000 / mth", ""),
    ("CAEP hosting — Months 4-6",            "R 30 000 / mth", ""),
    ("CAEP hosting — Month 7 onwards",       "R 40 000 / mth", "Steady-state representative hosting fee"),
    ("CAEP AUM fee",                         "0.25% p.a.",     "Annual fee on total AUM paid to CAEP; modelled as monthly accrual (÷12)"),
    ("Software & infrastructure",            "R 5 000 / mth",  "Cloud, Supabase, VALR API, domain, backups"),
    ("Company secretarial (CIPC etc.)",      "R 3 000 p.a.",   "Billed in months 12, 24 and 36 of the model"),
    ("Founder salary / drawings",            "R 0",            "Excluded — model shows pure company breakeven on operating revenue"),
    ("Simon Hobday capital",                 "Excluded",       "R500k equity + R500k loan excluded; model shows operating-revenue breakeven"),
])

row = asm_block(ws_a, row, "CLIENT ASSUMPTIONS", [
    ("Client 1 — initial lump sum",          "R 300 000",      "Existing signed-up client; Month 1"),
    ("Client 1 — monthly contribution",      "R 200 000 / mth","Minimum 5 years; starts Month 1"),
    ("Client 2 — monthly contribution",      "R 40 000 / mth", "Lined-up client; Month 2 start, no lump sum"),
    ("Client 3 — initial lump sum",          "R 5 000 000",    "Lined-up client; Month 2, no monthly contributions"),
    ("New client acquisition",               "1 per quarter",  "1 new R 50k/month client from Month 4, every 3 months"),
    ("New client monthly contribution",      "R 50 000 / mth", "Conservative flat assumption; no lump sum for new clients"),
])

row = asm_block(ws_a, row, "MODEL SUMMARY (calculated)", [
    ("Forecast horizon",                     "36 months",              "June 2026 — May 2029"),
    ("AUM at Month 36",                      f"R {m36_aum:,.0f}",      "End-of-month portfolio value after fees"),
    ("First month with positive net cash",   f"Month {first_pos_net} ({mo_label(first_pos_net) if first_pos_net else 'N/A'})", "Monthly revenue exceeds monthly costs"),
    ("Cumulative cash breakeven",            f"Month {first_pos_cum} ({mo_label(first_pos_cum) if first_pos_cum else 'N/A'})", "Cumulative cash turns positive"),
    ("Peak capital required (worst deficit)", f"R {abs(peak_deficit):,.0f}", "Maximum cumulative cash deficit before breakeven"),
], title_color=GOLD)


# ───────────────────────────────────────────────────────────
# SHEET 2: Monthly Forecast (main P&L / cashflow table)
# ───────────────────────────────────────────────────────────
ws_f = wb.create_sheet("Monthly Forecast")
ws_f.sheet_view.showGridLines = False
ws_f.freeze_panes = "C4"   # freeze month/label columns

# Column layout
# A  Month #    B  Label
# --- REVENUE ---
# C  Total Contributions   D  Platform Fee   E  Performance Fee   F  Total Revenue
# --- COSTS ---
# G  CAEP Onboard   H  CAEP Hosting   I  CAEP AUM Fee   J  Software   K  Co Sec   L  Total Costs
# --- CASH ---
# M  Net Cash   N  Cumulative Cash   O  Total AUM

COL_WIDTHS = [5, 9, 16, 14, 16, 14,   16, 14, 14, 11, 11, 13,   13, 16, 16]
for i, w in enumerate(COL_WIDTHS, 1):
    ws_f.column_dimensions[get_column_letter(i)].width = w

# Row 1: Sheet title
ws_f.merge_cells("A1:O1")
c = ws_f.cell(row=1, column=1,
    value="BitWealth (Pty) Ltd — 36-Month Operating Cashflow Forecast (June 2026 – May 2029)")
c.fill = fp(GOLD); c.font = Font(bold=True, sz=14, color=WHITE, name="Calibri")
c.alignment = aln(h="center")
ws_f.row_dimensions[1].height = 25

# Row 2: Section banners
section_header(ws_f, 2, 1, 2,  "PERIOD",    fh=DARK)
section_header(ws_f, 2, 3, 6,  "REVENUE",   fh="1B5E20")
section_header(ws_f, 2, 7, 12, "COSTS",     fh="B71C1C")
section_header(ws_f, 2, 13,15, "CASHFLOW",  fh=TEAL)

# Row 3: Column headers
HDR3 = [
    "Month", "Month",
    "Total Contributions", "Platform Fee\n(Revenue)", "Perf Fee\n(Revenue)", "TOTAL\nREVENUE",
    "CAEP\nOnboarding", "CAEP\nHosting", "CAEP AUM\nFee (cost)", "Software", "Co. Sec.",
    "TOTAL\nCOSTS",
    "NET MONTHLY\nCASH", "CUMULATIVE\nCASH", "TOTAL AUM\n(end of month)",
]
ws_f.row_dimensions[3].height = 36
for ci, h in enumerate(HDR3, 1):
    fh = "1B5E20" if 3 <= ci <= 6 else ("B71C1C" if 7 <= ci <= 12 else (TEAL if ci >= 13 else DARK))
    sc(ws_f, 3, ci, val=h, fh=fh, bold=True, sz=9, fc=WHITE, ha="center", wrap=True)

# Rows 4+: Data
for ri, r in enumerate(rows):
    row_num = ri + 4
    m = r['month']
    is_breakeven = (first_pos_cum is not None and m == first_pos_cum)
    row_bg = L_GOLD if is_breakeven else (L_GREY if m % 2 == 0 else WHITE)

    sc(ws_f, row_num, 1, val=m,       fh=row_bg, bold=is_breakeven, ha="center", border=bdr_all)
    sc(ws_f, row_num, 2, val=r['label'], fh=row_bg, bold=is_breakeven, ha="center", border=bdr_all)

    # Revenue
    mc(ws_f, row_num, 3,  r['contrib'],  fh=row_bg)
    mc(ws_f, row_num, 4,  r['plat'],     fh=row_bg)
    mc(ws_f, row_num, 5,  r['perf'],     fh=row_bg)
    mc(ws_f, row_num, 6,  r['rev'],      fh="E8F5E9", bold=True)

    # Costs
    mc(ws_f, row_num, 7,  r['c_onboard'], fh=row_bg)
    mc(ws_f, row_num, 8,  r['c_host'],    fh=row_bg)
    mc(ws_f, row_num, 9,  r['c_aum'],     fh=row_bg)
    mc(ws_f, row_num, 10, r['c_sw'],      fh=row_bg)
    mc(ws_f, row_num, 11, r['c_sec'],     fh=row_bg)
    mc(ws_f, row_num, 12, r['cost'],      fh="FFEBEE", bold=True)

    # Cashflow (sign-coloured)
    mc(ws_f, row_num, 13, r['net'],  bold=True, sign_color=True)
    mc(ws_f, row_num, 14, r['cum'],  bold=True, sign_color=True)
    mc(ws_f, row_num, 15, r['aum'],  fh=row_bg)

    if is_breakeven:
        # Add note in AUM column cell
        ws_f.cell(row=row_num, column=15).comment = None

# Totals row
tr = len(rows) + 4
sc(ws_f, tr, 1, val="TOTAL", fh=GOLD, bold=True, fc=WHITE, border=bdr_all)
ws_f.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
for ci, key in [(3,'contrib'),(4,'plat'),(5,'perf'),(6,'rev'),
                (7,'c_onboard'),(8,'c_host'),(9,'c_aum'),(10,'c_sw'),(11,'c_sec'),(12,'cost'),
                (13,'net')]:
    mc(ws_f, tr, ci, sum(r[key] for r in rows), fh=L_GOLD, bold=True)
mc(ws_f, tr, 14, rows[-1]['cum'],  fh=L_GOLD, bold=True)   # final cumulative
mc(ws_f, tr, 15, rows[-1]['aum'],  fh=L_GOLD, bold=True)   # final AUM

# Breakeven annotation
if first_pos_cum:
    br = first_pos_cum + 3  # row in sheet
    note_row = len(rows) + 6
    ws_f.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=15)
    c = ws_f.cell(row=note_row, column=1,
        value=f"★  CUMULATIVE BREAKEVEN: Month {first_pos_cum} ({mo_label(first_pos_cum)})  —  "
              f"Monthly net-positive first reached: Month {first_pos_net} ({mo_label(first_pos_net) if first_pos_net else 'N/A'})  —  "
              f"Peak capital required: R {abs(peak_deficit):,.0f}  —  "
              f"AUM at Month 36: R {m36_aum:,.0f}")
    c.fill = fp(L_GOLD); c.font = Font(bold=True, sz=10, color=GOLD, name="Calibri")
    c.alignment = aln(h="center")


# ── Cumulative Cashflow Line Chart ────────────────────────
chart_row_start = len(rows) + 8

chart = LineChart()
chart.title = "Cumulative Operating Cashflow"
chart.style = 10
chart.y_axis.title = "ZAR"
chart.x_axis.title = "Month"
chart.height = 12
chart.width  = 25

# Cumulative cash data (col N = 14)
data_ref = Reference(ws_f, min_col=14, min_row=3, max_row=len(rows)+3)
chart.add_data(data_ref, titles_from_data=True)
chart.series[0].graphicalProperties.line.solidFill = GOLD
chart.series[0].graphicalProperties.line.width     = 20000

# Revenue data (col F = 6)
rev_ref = Reference(ws_f, min_col=6, min_row=3, max_row=len(rows)+3)
chart.add_data(rev_ref, titles_from_data=True)
chart.series[1].graphicalProperties.line.solidFill = "1B5E20"
chart.series[1].graphicalProperties.line.width     = 15000

# Cost data (col L = 12)
cost_ref = Reference(ws_f, min_col=12, min_row=3, max_row=len(rows)+3)
chart.add_data(cost_ref, titles_from_data=True)
chart.series[2].graphicalProperties.line.solidFill = "B71C1C"
chart.series[2].graphicalProperties.line.width     = 15000

# Month labels
cat_ref = Reference(ws_f, min_col=2, min_row=4, max_row=len(rows)+3)
chart.set_categories(cat_ref)

ws_f.add_chart(chart, f"A{chart_row_start}")


# ───────────────────────────────────────────────────────────
# SHEET 3: AUM chart (AUM growth bar chart)
# ───────────────────────────────────────────────────────────
ws_aum = wb.create_sheet("AUM Growth")
ws_aum.sheet_view.showGridLines = False

ws_aum.merge_cells("A1:F1")
c = ws_aum.cell(row=1, column=1, value="BitWealth (Pty) Ltd — Total AUM Projection")
c.fill = fp(GOLD); c.font = Font(bold=True, sz=14, color=WHITE, name="Calibri")
c.alignment = aln(h="center")

# Write AUM data for chart
ws_aum.cell(row=3, column=1, value="Month")
ws_aum.cell(row=3, column=2, value="Month Label")
ws_aum.cell(row=3, column=3, value="Total AUM")
ws_aum.cell(row=3, column=4, value="Total Revenue")
ws_aum.cell(row=3, column=5, value="Total Costs")
ws_aum.cell(row=3, column=6, value="Net Monthly Cash")

for ci in range(1, 7):
    c = ws_aum.cell(row=3, column=ci)
    c.fill = fp(GOLD); c.font = Font(bold=True, sz=10, color=WHITE, name="Calibri")
    c.alignment = aln(h="center")

for ri, r in enumerate(rows, 4):
    ws_aum.cell(row=ri, column=1, value=r['month'])
    ws_aum.cell(row=ri, column=2, value=r['label'])
    ws_aum.cell(row=ri, column=3, value=round(r['aum'],    0)); ws_aum.cell(row=ri, column=3).number_format = MONEY_FMT
    ws_aum.cell(row=ri, column=4, value=round(r['rev'],    0)); ws_aum.cell(row=ri, column=4).number_format = MONEY_FMT
    ws_aum.cell(row=ri, column=5, value=round(r['cost'],   0)); ws_aum.cell(row=ri, column=5).number_format = MONEY_FMT
    ws_aum.cell(row=ri, column=6, value=round(r['net'],    0)); ws_aum.cell(row=ri, column=6).number_format = MONEY_FMT

for col, w in zip("ABCDEF", [7,10,18,16,16,18]):
    ws_aum.column_dimensions[col].width = w

# AUM Bar Chart
aum_chart = BarChart()
aum_chart.type = "col"
aum_chart.title = "Total AUM Growth (ZAR)"
aum_chart.style = 10
aum_chart.y_axis.title = "ZAR"
aum_chart.x_axis.title = "Month"
aum_chart.height = 14
aum_chart.width  = 28
aum_data = Reference(ws_aum, min_col=3, min_row=3, max_row=len(rows)+3)
aum_chart.add_data(aum_data, titles_from_data=True)
aum_chart.series[0].graphicalProperties.solidFill = GOLD
cats = Reference(ws_aum, min_col=2, min_row=4, max_row=len(rows)+3)
aum_chart.set_categories(cats)
ws_aum.add_chart(aum_chart, "A42")


# ───────────────────────────────────────────────────────────
# SHEET 4: Client Detail (per-client AUM + performance fees)
# ───────────────────────────────────────────────────────────
ws_c = wb.create_sheet("Client Detail")
ws_c.sheet_view.showGridLines = False

ws_c.merge_cells("A1:D1")
c = ws_c.cell(row=1, column=1, value="BitWealth — Per-Client Monthly AUM & Performance Fee Detail")
c.fill = fp(GOLD); c.font = Font(bold=True, sz=13, color=WHITE, name="Calibri")
c.alignment = aln(h="center")
ws_c.row_dimensions[1].height = 25

# How many clients do we have?
n_clients = len(all_clients)
# Columns: A=Month, B=Label, then per client: contrib, AUM, perf_fee (3 cols each)
COL_CLIENT_BASE = 3   # first client starts at column 3

ws_c.column_dimensions["A"].width = 7
ws_c.column_dimensions["B"].width = 9

# Row 2: client name headers (merged across 3 cols each)
for ci, client in enumerate(all_clients):
    col_start = COL_CLIENT_BASE + ci * 3
    ws_c.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start+2)
    c = ws_c.cell(row=2, column=col_start, value=client.name)
    c.fill = fp(TEAL); c.font = Font(bold=True, sz=9, color=WHITE, name="Calibri")
    c.alignment = aln(h="center")
    for col_off, w in enumerate([14, 16, 14]):
        ws_c.column_dimensions[get_column_letter(col_start + col_off)].width = w

# Row 3: sub-headers
sc(ws_c, 3, 1, "Month",  fh=DARK, fc=WHITE, bold=True, sz=9)
sc(ws_c, 3, 2, "Label",  fh=DARK, fc=WHITE, bold=True, sz=9)
for ci in range(n_clients):
    base = COL_CLIENT_BASE + ci * 3
    sc(ws_c, 3, base,   "Contribution",    fh="1B5E20", fc=WHITE, bold=True, sz=9)
    sc(ws_c, 3, base+1, "Closing AUM",     fh="1B5E20", fc=WHITE, bold=True, sz=9)
    sc(ws_c, 3, base+2, "Perf Fee (crys.)",fh="1B5E20", fc=WHITE, bold=True, sz=9)

# Data rows
for ri, r in enumerate(rows, 4):
    m   = r['month']
    bg  = L_GREY if m % 2 == 0 else WHITE
    sc(ws_c, ri, 1, val=m,         fh=bg, ha="center", border=bdr_all)
    sc(ws_c, ri, 2, val=r['label'],fh=bg, ha="center", border=bdr_all)
    for ci, cd in enumerate(r['cd']):
        base = COL_CLIENT_BASE + ci * 3
        mc(ws_c, ri, base,   cd['contrib'] if cd['contrib'] > 0 else None, fh=bg)
        mc(ws_c, ri, base+1, cd['closing'] if cd['closing'] > 0 else None, fh=bg)
        mc(ws_c, ri, base+2, cd['perf']    if cd['perf']    > 0 else None, fh=bg if cd['perf'] == 0 else L_GREEN)

# Totals
tr2 = len(rows) + 4
sc(ws_c, tr2, 1, "TOTAL", fh=L_GOLD, bold=True, border=bdr_all)
sc(ws_c, tr2, 2, "",      fh=L_GOLD, border=bdr_all)
for ci in range(n_clients):
    base = COL_CLIENT_BASE + ci * 3
    mc(ws_c, tr2, base,   sum(r['cd'][ci]['contrib'] for r in rows), fh=L_GOLD, bold=True)
    mc(ws_c, tr2, base+1, rows[-1]['cd'][ci]['closing'],              fh=L_GOLD, bold=True)
    mc(ws_c, tr2, base+2, sum(r['cd'][ci]['perf'] for r in rows),    fh=L_GOLD, bold=True)


# ═══════════════════════════════════════════════════════════
# 5.  SAVE
# ═══════════════════════════════════════════════════════════
out_path = OUT_DIR / "BitWealth_Cashflow_Forecast_v1.0.xlsx"
wb.save(out_path)
print(f"\nSaved: {out_path}")
