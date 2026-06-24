"""
BitWealth (Pty) Ltd — 96-Month Operating Cashflow Forecast v4.0
Generates TWO workbooks in a single run:
  · BitWealth_Cashflow_Forecast_v4.0.xlsx         — CAEP as FSP (fixed cost)
  · BitWealth_Cashflow_Forecast_v4.0_Finova.xlsx  — Finova Capital as FSP (rev-share)

Changes from v3:
  · 96 months (8 years) — extended from 48
  · 3-tier return: 60% p.a. Y1-3  →  10% Y4  →  adjustable Y5-8 (default 50%)
    All three rates are independent Variables; Y5-8 defaults to 50% (next BTC cycle)
  · Exchange fee share revenue added (VALR deal: 50% of 0.1% fee on contributions)
    Covers ZAR/USDT, USDT/BTC, and USDT/USDPC conversion fees
  · Finova Capital variant: 20% perf-fee share + 50% platform-fee share, no fixed FSP costs
  · All v3 variable values carried across

Column layout (40 cols):
  A-C  (1-3)   Period
  D-X  (4-24)  Client detail — hidden, grouped (click + to expand)
  Y-AC (25-29) Revenue: Total Contrib | Plat Fee | Perf Fee | Exchange Fee Rev | TOTAL
  AD   (30)    Total AUM
  AE-AJ(31-36) Costs: FSP1 | FSP2 | FSP3 | Software | Co Sec | TOTAL
  AK   (37)    Reserve Provision
  AL   (38)    Net Cash
  AM   (39)    Cumulative Cash
  AN   (40)    Reserve Fund Balance
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter as gcl

OUT_DIR = Path("docs") / "Financial"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════
# STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════
GOLD    = "C9A227"; L_GOLD  = "F5EBC8"; DARK    = "1A1A1A"; WHITE   = "FFFFFF"
YELLOW  = "FFF9C4"; L_GREEN = "C8E6C9"; L_RED   = "FFCDD2"; L_GREY  = "F5F5F5"
L_BLUE  = "E3F2FD"; TEAL    = "006064"; GREEN_D = "1B5E20"; RED_D   = "B71C1C"
ORANGE  = "E65100"; L_ORANGE = "FFF3E0"
MONEY   = "R #,##0;[Red](R #,##0)"; PCT = "0.00%"

# Row constants (shared between build_forecast and build_charts)
HDR_ROW    = 3
DATA_START = 4
MAX_ROWS   = 96


# ═══════════════════════════════════════════════════════════════
# STYLE HELPERS
# ═══════════════════════════════════════════════════════════════
def fp(h):  return PatternFill("solid", fgColor=h)
def fn(bold=False, sz=11, col=DARK, ital=False):
    return Font(bold=bold, size=sz, color=col, italic=ital, name="Calibri")
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_t = Side(border_style="thin",   color="CCCCCC")
_T = Side(border_style="medium", color="888888")
bdr_all = Border(left=_t, right=_t, top=_t, bottom=_t)
bdr_hdr = Border(left=_t, right=_t, top=_T, bottom=_T)


def sc(ws, r, c, val=None, fh=None, bold=False, sz=11, fc=DARK,
       ital=False, ha="center", nf=None, wrap=False, border=None):
    cell = ws.cell(row=r, column=c, value=val)
    if fh:     cell.fill      = fp(fh)
    cell.font      = fn(bold=bold, sz=sz, col=fc, ital=ital)
    cell.alignment = al(h=ha, v="center", wrap=wrap)
    if nf:     cell.number_format = nf
    if border: cell.border    = border
    return cell


def section_hdr(ws, row, c1, c2, text, fh=TEAL):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill = fp(fh)
    c.font = Font(bold=True, sz=10, color=WHITE, name="Calibri")
    c.alignment = al(h="center")


# ═══════════════════════════════════════════════════════════════════════════
# BUILD VARIABLES SHEET
# ═══════════════════════════════════════════════════════════════════════════
def build_variables(ws_v, mode):
    """Populate the VARIABLES sheet.  Returns the VARS dict of cell references."""
    ws_v.sheet_view.showGridLines = False
    ws_v.column_dimensions["A"].width = 55
    ws_v.column_dimensions["B"].width = 22
    ws_v.column_dimensions["C"].width = 45

    title_bg = GOLD if mode == "caep" else ORANGE
    tag = "" if mode == "caep" else "  [FINOVA CAPITAL VARIANT]"

    ws_v.merge_cells("A1:C1")
    c = ws_v.cell(row=1, column=1,
        value=f"BitWealth (Pty) Ltd — Cashflow Model Variables{tag}")
    c.fill = fp(title_bg); c.font = Font(bold=True, sz=16, color=WHITE, name="Calibri")
    c.alignment = al(h="center"); ws_v.row_dimensions[1].height = 30

    ws_v.merge_cells("A2:C2")
    fsp_label = "CAEP (fixed cost)" if mode == "caep" else "FINOVA CAPITAL (rev-share)"
    c = ws_v.cell(row=2, column=1,
        value=(f"v4.0  ·  96-Month Forecast (8 Years)  ·  Adjust YELLOW cells  ·  "
               f"Forecast tab auto-recalculates  ·  FSP: {fsp_label}"))
    c.fill = fp(L_GOLD); c.font = fn(sz=10, ital=True); c.alignment = al(h="center")

    V = {}  # key -> "VARIABLES!$B$N"

    def vr(start_row, title_text, items, title_color=TEAL):
        """Write a variable block and register cell refs in V. Returns next free row."""
        r = start_row
        ws_v.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ch = ws_v.cell(row=r, column=1, value=title_text)
        ch.fill = fp(title_color)
        ch.font = Font(bold=True, sz=11, color=WHITE, name="Calibri")
        ch.alignment = al(h="left"); ws_v.row_dimensions[r].height = 18
        r += 1
        for key, label, value, note, editable in items:
            sc(ws_v, r, 1, val=label, ha="left", fh=L_GREY, border=bdr_all)
            cb = ws_v.cell(row=r, column=2, value=value)
            cb.fill = fp(YELLOW if editable else WHITE)
            cb.font = fn(bold=editable, sz=11)
            cb.alignment = al(h="right")
            cb.border = bdr_hdr if editable else bdr_all
            if isinstance(value, float) and 0 < value < 1:
                cb.number_format = PCT
            elif isinstance(value, (int, float)) and value >= 1000:
                cb.number_format = MONEY
            sc(ws_v, r, 3, val=note, ha="left", fh=WHITE,
               ital=True, sz=9, fc="555555", border=bdr_all)
            V[key] = f"VARIABLES!$B${r}"
            r += 1
        return r + 1  # blank gap after block

    # ── RETURN & PERFORMANCE ────────────────────────────────────────────
    row = vr(4, "RETURN & PERFORMANCE ASSUMPTIONS", [
        ("ret_y1_3",   "Annual portfolio return — Years 1–3 (p.a.)",           0.60,
         "60% p.a. BTC bull cycle (months 1 → y4_start-1); divided by 12 for monthly", True),
        ("ret_y4",     "Annual portfolio return — Year 4 (p.a.)",               0.10,
         "10% moderate/consolidation (months y4_start → y5_start-1)", True),
        ("y4_start",   "Year 4 starts (model month #)",                          37,
         "Month 37 = July 2029", True),
        ("y5_start",   "Year 5 starts (model month #)",                          49,
         "Month 49 = July 2030 — start of next BTC halving cycle", True),
        ("ret_y5_8",   "Annual portfolio return — Years 5–8 (p.a.)",             0.50,
         "Next BTC bull cycle. Set to 0% for conservative Y5-8 assumption", True),
        ("perf_period","Performance fee crystallisation period (months)",         12,
         "12 = annual crystallisation (simplified to monthly accrual in formulas)", True),
        ("reserve_pm", "Monthly liquidity buffer provision (R)",                  10580,
         "Set aside each month in Y1-3 to fund cash deficit in Y4; tracked in Reserve Fund Bal", True),
    ])

    # ── EXCHANGE FEE SHARE ──────────────────────────────────────────────
    row = vr(row, "EXCHANGE FEE SHARE REVENUE  (VALR revenue-share deal)", [
        ("exch_fee_rate",  "VALR exchange fee rate per trade",                    0.001,
         "0.1% default. Covers ZAR/USDT, USDT/BTC, and USDT/USDPC conversions. Adjust if VALR rate changes", True),
        ("exch_fee_share", "BitWealth share of exchange fees",                    0.50,
         "50% — as negotiated with VALR. Revenue = total monthly contributions × fee_rate × this share", True),
    ], title_color=GREEN_D)

    # ── FSP COSTS (mode-dependent) ──────────────────────────────────────
    if mode == "caep":
        row = vr(row, "FSP COSTS — CAEP (fixed cost model)", [
            ("caep_ob",  "CAEP onboarding fee (once-off, Month 1)",               57500,
             "Juristic Representative appointment; non-recurring", False),
            ("caep_h13", "CAEP hosting — Months 1–3 (per month)",                 23000, "", True),
            ("caep_h46", "CAEP hosting — Months 4–6 (per month)",                 34500, "", True),
            ("caep_h7p", "CAEP hosting — Month 7+ (per month, steady-state)",     46000,
             "Negotiate a contractual cap; dominant cost driver", True),
            ("caep_aum", "CAEP AUM fee (% of AUM, billed annually)",               0.0025,
             "0.25% p.a.; charged in months 12, 24, 36 … 96 on end-of-month AUM", True),
            ("software", "Software & infrastructure (monthly)",                    5000,
             "Cloud, Supabase, VALR API, domain etc.", True),
            ("co_sec",   "Company secretarial (annual, months 12/24/36 …)",        3000,
             "CIPC annual return + company secretary", True),
        ], title_color=RED_D)
    else:  # finova
        row = vr(row, "FSP ARRANGEMENT — FINOVA CAPITAL (rev-share model)", [
            ("finova_perf_share", "Performance fee share paid to Finova Capital",  0.20,
             "20% of all performance fees earned by BitWealth. BitWealth retains 80%", True),
            ("finova_plat_share", "Platform fee share paid to Finova Capital",     0.50,
             "50% of all platform fees earned by BitWealth. BitWealth retains 50%", True),
            ("software",          "Software & infrastructure (monthly)",            5000,
             "Cloud, Supabase, VALR API, domain etc.", True),
            ("co_sec",            "Company secretarial (annual, months 12/24/36 …)", 3000,
             "CIPC annual return + company secretary", True),
        ], title_color=ORANGE)

    # ── CLIENTS 1–4 ─────────────────────────────────────────────────────
    row = vr(row, "CLIENT 1 (Founding — preferential 2.5% perf fee)", [
        ("c1_start",   "Start month",             1,      "Month 1 = June 2026", True),
        ("c1_lump",    "Initial lump sum (R)",     300000, "", True),
        ("c1_monthly", "Monthly contribution (R)", 200000, "", True),
        ("c1_plat",    "Platform fee rate",        0.0075, "0.75%", True),
        ("c1_perf",    "Performance fee rate",     0.025,  "2.5% preferential", True),
    ], title_color="37474F")

    row = vr(row, "CLIENT 2 (Founding)", [
        ("c2_start",   "Start month",             2,     "Month 2 = July 2026", True),
        ("c2_lump",    "Initial lump sum (R)",     0,     "", True),
        ("c2_monthly", "Monthly contribution (R)", 40000, "", True),
        ("c2_plat",    "Platform fee rate",        0.0075,"", True),
        ("c2_perf",    "Performance fee rate",     0.10,  "Standard 10%", True),
    ], title_color="37474F")

    row = vr(row, "CLIENT 3 (Founding — large lump sum, no monthly contributions)", [
        ("c3_start",   "Start month",             2,       "Month 2 = July 2026", True),
        ("c3_lump",    "Initial lump sum (R)",     5000000, "0.75% platform fee = R37 500 on day 1", True),
        ("c3_monthly", "Monthly contribution (R)", 0,       "No monthly contributions", True),
        ("c3_plat",    "Platform fee rate",        0.0075,  "", True),
        ("c3_perf",    "Performance fee rate",     0.10,    "", True),
    ], title_color="37474F")

    row = vr(row, "CLIENT 4 (Founding — preferential 2.5% perf fee)", [
        ("c4_start",   "Start month",             2,     "Month 2 = July 2026", True),
        ("c4_lump",    "Initial lump sum (R)",     0,     "", True),
        ("c4_monthly", "Monthly contribution (R)", 20000, "", True),
        ("c4_plat",    "Platform fee rate",        0.0075,"", True),
        ("c4_perf",    "Performance fee rate",     0.025, "2.5% preferential", True),
    ], title_color="37474F")

    # ── NEW CLIENTS ─────────────────────────────────────────────────────
    row = vr(row, "NEW CLIENT ACQUISITION (rolling monthly ramp)", [
        ("nc_max",       "Maximum cohort batches to model",                       48,
         "48 batches × interval=1 month → fills months 3-50; stays at 48 clients from month 50", True),
        ("nc_start",     "First cohort — start month",                            3,
         "Month 3 = August 2026", True),
        ("nc_interval",  "Cohort interval (months)",                              1,
         "1 = monthly. Change to 3 for quarterly onboarding", True),
        ("nc_per_cohort","Clients per cohort batch",                               1,
         "Change to 2+ to onboard multiple clients per interval; scales contributions & lump sums", True),
        ("nc_lump",      "New client lump sum (R, per client)",                   50000,
         "Per client on join month; multiply by nc_per_cohort for batch total", True),
        ("nc_monthly",   "New client monthly contribution (R, per client)",       5000,
         "Per client per month; total = active_clients × this value", True),
        ("nc_plat",      "New client platform fee rate",                           0.0075, "Standard 0.75%", True),
        ("nc_perf",      "New client performance fee rate",                        0.10,   "Standard 10%", True),
    ], title_color=GREEN_D)

    # ── FORECAST PARAMETERS ─────────────────────────────────────────────
    row = vr(row, "FORECAST PARAMETERS", [
        ("fc_months", "Forecast months",                            96,   "96 = 8 full years", True),
        ("fc_year",   "Model start year",                           2026, "Calendar year of Month 1", False),
        ("fc_month",  "Model start calendar month (1=Jan … 12=Dec)", 6,  "6 = June", False),
    ], title_color=GOLD)

    # ── Legend note ─────────────────────────────────────────────────────
    lr = row + 1
    ws_v.merge_cells(f"A{lr}:C{lr}")
    if mode == "caep":
        leg = ("★ Yellow cells are editable — Forecast tab auto-recalculates.  "
               "RETURNS: 3-tier: 60% p.a. Y1-3 → 10% Y4 → 50% Y5-8 (next BTC cycle; all adjustable).  "
               "EXCHANGE FEE: BitWealth earns 50% of VALR's 0.1% fee on each monthly contribution "
               "(ZAR/USDT + USDT/BTC + USDT/USDPC).  Adjust exch_fee_rate and exch_fee_share as needed.  "
               "PERF FEE: simplified monthly accrual (AUM × monthly_return × perf_rate); "
               "approximates annual crystallisation.  CAEP AUM fee: 0.25% billed annually on end-of-month AUM.")
    else:
        leg = ("★ FINOVA CAPITAL VARIANT — compare side-by-side with the CAEP version.  "
               "Finova takes 20% of all performance fees + 50% of all platform fees (no fixed hosting costs).  "
               "Advantage: lower risk in early months (no large fixed bill).  "
               "Disadvantage: revenue share escalates as AUM and fees grow — CAEP may become cheaper above a "
               "certain AUM threshold.  All client, return, and exchange-fee assumptions are identical to the CAEP file.")
    c = ws_v.cell(row=lr, column=1, value=leg)
    c.fill = fp(L_GOLD if mode == "caep" else L_ORANGE)
    c.font = fn(sz=9, ital=True)
    c.alignment = al(h="left", wrap=True)
    ws_v.row_dimensions[lr].height = 60

    return V


# ═══════════════════════════════════════════════════════════════════════════
# BUILD FORECAST SHEET
# ═══════════════════════════════════════════════════════════════════════════
def build_forecast(ws_f, V, mode):
    ws_f.sheet_view.showGridLines = False

    # Column widths
    cw = {
        1: 6,  2: 9,  3: 8,
        4:13,  5:12,  6:12,  7:15,
        8:13,  9:12, 10:12, 11:15,
       12:13, 13:12, 14:12, 15:15,
       16:13, 17:12, 18:12, 19:15,
       20: 8, 21:14, 22:12, 23:12, 24:15,
       25:14, 26:14, 27:14, 28:14, 29:15,
       30:16,
       31:14, 32:14, 33:14, 34:11, 35:11, 36:13,
       37:14, 38:15, 39:16, 40:16,
    }
    for ci, w in cw.items():
        ws_f.column_dimensions[gcl(ci)].width = w

    # Group and hide detail columns D:X
    for ci in range(4, 25):
        ws_f.column_dimensions[gcl(ci)].outlineLevel = 1
        ws_f.column_dimensions[gcl(ci)].hidden = True
    ws_f.sheet_format.outlineLevelCol = 1
    ws_f.freeze_panes = f"C{DATA_START}"

    # Title row
    fsp_label = "CAEP" if mode == "caep" else "FINOVA CAPITAL"
    title_bg  = GOLD   if mode == "caep" else ORANGE
    ws_f.merge_cells(f"A1:{gcl(40)}1")
    c = ws_f.cell(row=1, column=1,
        value=(f"BitWealth (Pty) Ltd — 96-Month Cashflow Forecast v4.0  ·  "
               f"FSP: {fsp_label}  ·  All formulas auto-update when VARIABLES sheet changes"))
    c.fill = fp(title_bg); c.font = Font(bold=True, sz=13, color=WHITE, name="Calibri")
    c.alignment = al(h="center"); ws_f.row_dimensions[1].height = 22

    # Section banners (row 2)
    cost_bg = RED_D if mode == "caep" else ORANGE
    section_hdr(ws_f, 2,  1,  3, "PERIOD",                               fh=DARK)
    section_hdr(ws_f, 2,  4, 24, "CLIENT DETAIL  (hidden — click + to expand)", fh="37474F")
    section_hdr(ws_f, 2, 25, 29, "REVENUE",                              fh=GREEN_D)
    section_hdr(ws_f, 2, 30, 30, "TOTAL AUM",                            fh=TEAL)
    section_hdr(ws_f, 2, 31, 36, "COSTS",                                fh=cost_bg)
    section_hdr(ws_f, 2, 37, 37, "RESERVE",                              fh="6D4C41")
    section_hdr(ws_f, 2, 38, 40, "CASHFLOW",                             fh=TEAL)

    # Column headers (row 3)
    ws_f.row_dimensions[HDR_ROW].height = 52
    fsp_h = {
        "caep":   {31: "CAEP\nOnboard",    32: "CAEP\nHosting",       33: "CAEP\nAUM Fee"},
        "finova": {31: "Finova\nSetup\n(R0)", 32: "Finova\nPlat\nShare", 33: "Finova\nPerf\nShare"},
    }[mode]
    hdrs = [
        (1,"Month\n#"), (2,"Month"), (3,"Monthly\nReturn %"),
        (4,"C1\nContrib"), (5,"C1\nPlat"), (6,"C1\nPerf"), (7,"C1\nAUM"),
        (8,"C2\nContrib"), (9,"C2\nPlat"),(10,"C2\nPerf"),(11,"C2\nAUM"),
       (12,"C3\nContrib"),(13,"C3\nPlat"),(14,"C3\nPerf"),(15,"C3\nAUM"),
       (16,"C4\nContrib"),(17,"C4\nPlat"),(18,"C4\nPerf"),(19,"C4\nAUM"),
       (20,"NC\nClients"),(21,"NC\nContrib"),(22,"NC\nPlat"),(23,"NC\nPerf"),(24,"NC\nAUM"),
       (25,"Total\nContrib"),(26,"Total\nPlat Fee"),(27,"Total\nPerf Fee"),
       (28,"Exchange\nFee Rev"),(29,"TOTAL\nREVENUE"),
       (30,"TOTAL\nAUM"),
       (31, fsp_h[31]),(32, fsp_h[32]),(33, fsp_h[33]),(34,"Software"),(35,"Co Sec"),
       (36,"TOTAL\nCOSTS"),
       (37,"Reserve\nProvision"),
       (38,"NET CASH\n(after reserve)"),(39,"CUMULATIVE\nCASH"),(40,"Reserve\nFund Bal"),
    ]
    for ci, text in hdrs:
        if   4 <= ci <= 24:             fh_h = "37474F"
        elif 25 <= ci <= 29:            fh_h = GREEN_D
        elif ci == 30:                  fh_h = TEAL
        elif 31 <= ci <= 36:            fh_h = cost_bg
        elif ci == 37:                  fh_h = "6D4C41"
        else:                           fh_h = TEAL
        sc(ws_f, HDR_ROW, ci, val=text, fh=fh_h, bold=True, sz=9,
           fc=WHITE, ha="center", wrap=True)

    # ────────────────────────────────────────────────────────────────────
    # FORMULA FACTORIES  (closures over V, mode)
    # ────────────────────────────────────────────────────────────────────
    def cl(n): return gcl(n)

    def f_label(r):
        return (f"=TEXT(DATE({V['fc_year']}+INT(({V['fc_month']}+$A{r}-2)/12),"
                f"MOD({V['fc_month']}+$A{r}-2,12)+1,1),\"mmm-yy\")")

    def f_return(r):
        """3-tier: Y1-3 / Y4 / Y5-8."""
        return (f"=IF($A{r}<{V['y4_start']},{V['ret_y1_3']}/12,"
                f"IF($A{r}<{V['y5_start']},{V['ret_y4']}/12,{V['ret_y5_8']}/12))")

    # Per-client helpers
    def f_contrib(r, start_v, lump_v, monthly_v):
        return f"=IF($A{r}<{start_v},0,IF($A{r}={start_v},{lump_v}+{monthly_v},{monthly_v}))"

    def f_platfee(r, contrib_col, plat_v):
        return f"={cl(contrib_col)}{r}*{plat_v}"

    def f_perffee(r, prev_aum_col, contrib_col, ret_col, perf_v, start_v):
        prev = f"IF($A{r}={start_v},0,{cl(prev_aum_col)}{r-1})"
        return (f"=IF($A{r}<{start_v},0,"
                f"({prev}+{cl(contrib_col)}{r})*{cl(ret_col)}{r}*{perf_v})")

    def f_closing_aum(r, prev_aum_col, contrib_col, ret_col, perf_col, start_v):
        prev = f"IF($A{r}={start_v},0,{cl(prev_aum_col)}{r-1})"
        return (f"=IF($A{r}<{start_v},0,"
                f"({prev}+{cl(contrib_col)}{r})*(1+{cl(ret_col)}{r})-{cl(perf_col)}{r})")

    # New-client pool helpers
    def f_nc_count(r):
        batches = (f"MIN({V['nc_max']},"
                   f"INT(($A{r}-{V['nc_start']})/{V['nc_interval']})+1)")
        return f"=IF($A{r}<{V['nc_start']},0,{batches}*{V['nc_per_cohort']})"

    def f_nc_contrib(r, count_col):
        new_batch = (f"IF(AND($A{r}>={V['nc_start']},"
                     f"MOD($A{r}-{V['nc_start']},{V['nc_interval']})=0,"
                     f"INT(($A{r}-{V['nc_start']})/{V['nc_interval']})<{V['nc_max']}),"
                     f"{V['nc_per_cohort']}*{V['nc_lump']},0)")
        return f"={cl(count_col)}{r}*{V['nc_monthly']}+{new_batch}"

    def f_nc_platfee(r, contrib_col):
        return f"={cl(contrib_col)}{r}*{V['nc_plat']}"

    def f_nc_perffee(r, prev_aum_col, contrib_col, ret_col):
        prev = f"IF($A{r}<{V['nc_start']}+1,0,{cl(prev_aum_col)}{r-1})"
        return (f"=IF($A{r}<{V['nc_start']},0,"
                f"({prev}+{cl(contrib_col)}{r})*{cl(ret_col)}{r}*{V['nc_perf']})")

    def f_nc_closing_aum(r, prev_aum_col, contrib_col, ret_col, perf_col):
        prev = f"IF($A{r}<{V['nc_start']}+1,0,{cl(prev_aum_col)}{r-1})"
        return (f"=IF($A{r}<{V['nc_start']},0,"
                f"({prev}+{cl(contrib_col)}{r})*(1+{cl(ret_col)}{r})-{cl(perf_col)}{r})")

    # Revenue summary helpers
    def f_total_contrib(r):
        return (f"={cl(4)}{r}+{cl(8)}{r}+{cl(12)}{r}+{cl(16)}{r}+{cl(21)}{r}")

    def f_total_platfee(r):
        return (f"={cl(5)}{r}+{cl(9)}{r}+{cl(13)}{r}+{cl(17)}{r}+{cl(22)}{r}")

    def f_total_perffee(r):
        return (f"={cl(6)}{r}+{cl(10)}{r}+{cl(14)}{r}+{cl(18)}{r}+{cl(23)}{r}")

    def f_exch_fee_rev(r):
        # Revenue from VALR exchange-fee share on all monthly contributions
        return f"={cl(25)}{r}*{V['exch_fee_rate']}*{V['exch_fee_share']}"

    def f_total_rev(r):
        # Platform fees + Perf fees + Exchange fee share
        return f"={cl(26)}{r}+{cl(27)}{r}+{cl(28)}{r}"

    def f_total_aum(r):
        return (f"={cl(7)}{r}+{cl(11)}{r}+{cl(15)}{r}+{cl(19)}{r}+{cl(24)}{r}")

    # FSP cost helpers (mode-dependent)
    def f_fsp1(r):
        """CAEP: once-off onboarding.  Finova: R0 (no setup fee)."""
        if mode == "caep":
            return f"=IF($A{r}=1,{V['caep_ob']},0)"
        return "=0"

    def f_fsp2(r):
        """CAEP: tiered hosting.  Finova: platform-fee share paid to Finova."""
        if mode == "caep":
            return (f"=IF($A{r}<=3,{V['caep_h13']},"
                    f"IF($A{r}<=6,{V['caep_h46']},{V['caep_h7p']}))")
        # Finova: 50% of total platform fees
        return f"={cl(26)}{r}*{V['finova_plat_share']}"

    def f_fsp3(r):
        """CAEP: annual AUM fee.  Finova: perf-fee share paid to Finova."""
        if mode == "caep":
            # col 30 = Total AUM
            return f"=IF(MOD($A{r},12)=0,{cl(30)}{r}*{V['caep_aum']},0)"
        # Finova: 20% of total performance fees
        return f"={cl(27)}{r}*{V['finova_perf_share']}"

    def f_software(r):
        return f"={V['software']}"

    def f_co_sec(r):
        return f"=IF(MOD($A{r},12)=0,{V['co_sec']},0)"

    def f_total_costs(r):
        return (f"={cl(31)}{r}+{cl(32)}{r}+{cl(33)}{r}+"
                f"{cl(34)}{r}+{cl(35)}{r}")

    def f_reserve_prov(r):
        return f"=IF($A{r}<{V['y4_start']},{V['reserve_pm']},0)"

    def f_net_cash(r):
        # Total Revenue (29) - Total Costs (36) - Reserve (37)
        return f"={cl(29)}{r}-{cl(36)}{r}-{cl(37)}{r}"

    def f_cum_cash(r):
        if r == DATA_START:
            return f"={cl(38)}{r}"
        return f"={cl(39)}{r-1}+{cl(38)}{r}"

    def f_reserve_fund(r):
        if r == DATA_START:
            return f"={cl(37)}{r}"
        return f"={cl(40)}{r-1}+{cl(37)}{r}"

    # ────────────────────────────────────────────────────────────────────
    # WRITE DATA ROWS
    # ────────────────────────────────────────────────────────────────────
    def smf(col, formula, bg=None, bold=False, nf=MONEY, row=None):
        """Set a money/formula cell."""
        cell = ws_f.cell(row=row, column=col)
        cell.value = formula
        cell.font = fn(sz=10, bold=bold)
        cell.number_format = nf
        cell.alignment = al(h="right", v="center")
        if bg:
            cell.fill = fp(bg)
        cell.border = bdr_all

    for m in range(1, MAX_ROWS + 1):
        r = DATA_START + m - 1
        bg = L_GREY if m % 2 == 0 else WHITE

        # Month# — hardcoded integer (stable anchor for all IF($A{r}…) refs)
        sc(ws_f, r, 1, val=m, ha="center", fh=bg, border=bdr_all)

        # Month label
        c2 = ws_f.cell(row=r, column=2)
        c2.value = f_label(r); c2.fill = fp(bg); c2.font = fn(sz=10)
        c2.alignment = al(h="center"); c2.border = bdr_all

        # Monthly return rate
        c3 = ws_f.cell(row=r, column=3)
        c3.value = f_return(r); c3.fill = fp(bg)
        c3.font = fn(sz=10, ital=True, col="555555")
        c3.alignment = al(h="center"); c3.number_format = "0.00%"; c3.border = bdr_all

        # Partial so we can pass row
        def sm(col, formula, bgc=bg, bold=False):
            smf(col, formula, bgc, bold, MONEY, row=r)

        # Client 1
        sm(4,  f_contrib(r, V["c1_start"], V["c1_lump"], V["c1_monthly"]))
        sm(5,  f_platfee(r, 4,  V["c1_plat"]))
        sm(6,  f_perffee(r, 7,  4,  3, V["c1_perf"], V["c1_start"]))
        sm(7,  f_closing_aum(r, 7,  4,  3, 6,  V["c1_start"]))

        # Client 2
        sm(8,  f_contrib(r, V["c2_start"], V["c2_lump"], V["c2_monthly"]))
        sm(9,  f_platfee(r, 8,  V["c2_plat"]))
        sm(10, f_perffee(r, 11, 8,  3, V["c2_perf"], V["c2_start"]))
        sm(11, f_closing_aum(r, 11, 8,  3, 10, V["c2_start"]))

        # Client 3
        sm(12, f_contrib(r, V["c3_start"], V["c3_lump"], V["c3_monthly"]))
        sm(13, f_platfee(r, 12, V["c3_plat"]))
        sm(14, f_perffee(r, 15, 12, 3, V["c3_perf"], V["c3_start"]))
        sm(15, f_closing_aum(r, 15, 12, 3, 14, V["c3_start"]))

        # Client 4
        sm(16, f_contrib(r, V["c4_start"], V["c4_lump"], V["c4_monthly"]))
        sm(17, f_platfee(r, 16, V["c4_plat"]))
        sm(18, f_perffee(r, 19, 16, 3, V["c4_perf"], V["c4_start"]))
        sm(19, f_closing_aum(r, 19, 16, 3, 18, V["c4_start"]))

        # New clients — count (not currency)
        cc = ws_f.cell(row=r, column=20)
        cc.value = f_nc_count(r); cc.fill = fp(bg); cc.font = fn(sz=10)
        cc.number_format = "0"; cc.alignment = al(h="center"); cc.border = bdr_all

        sm(21, f_nc_contrib(r, 20))
        sm(22, f_nc_platfee(r, 21))
        sm(23, f_nc_perffee(r, 24, 21, 3))
        sm(24, f_nc_closing_aum(r, 24, 21, 3, 23))

        # Revenue summary
        sm(25, f_total_contrib(r))
        sm(26, f_total_platfee(r))
        sm(27, f_total_perffee(r))
        sm(28, f_exch_fee_rev(r))
        sm(29, f_total_rev(r),    bgc=L_GREEN, bold=True)

        # Total AUM
        sm(30, f_total_aum(r),    bgc=L_BLUE,  bold=True)

        # Costs
        sm(31, f_fsp1(r))
        sm(32, f_fsp2(r))
        sm(33, f_fsp3(r))
        sm(34, f_software(r))
        sm(35, f_co_sec(r))
        sm(36, f_total_costs(r),  bgc=L_RED,   bold=True)

        # Reserve
        sm(37, f_reserve_prov(r), bgc=L_ORANGE)

        # Cashflow
        sm(38, f_net_cash(r),     bold=True)
        sm(39, f_cum_cash(r),     bold=True)
        sm(40, f_reserve_fund(r), bgc=L_ORANGE, bold=True)

    # ── Totals row ──────────────────────────────────────────────────────
    TR = DATA_START + MAX_ROWS
    ws_f.merge_cells(start_row=TR, start_column=1, end_row=TR, end_column=2)
    ws_f.cell(row=TR, column=1,
        value=f"TOTAL / FINAL (96 months)").fill = fp(L_GOLD)
    ws_f.cell(row=TR, column=1).font      = fn(bold=True, sz=10)
    ws_f.cell(row=TR, column=1).alignment = al(h="center")

    DR = f"{DATA_START}:{DATA_START+MAX_ROWS-1}"  # row range string
    for ci in [25,26,27,28,29,31,32,33,34,35,36,37,38]:
        c = ws_f.cell(row=TR, column=ci)
        c.value = f"=SUM({gcl(ci)}{DATA_START}:{gcl(ci)}{DATA_START+MAX_ROWS-1})"
        c.fill = fp(L_GOLD); c.font = fn(bold=True, sz=10)
        c.number_format = MONEY; c.alignment = al(h="right"); c.border = bdr_all
    for ci in [30, 39, 40]:
        c = ws_f.cell(row=TR, column=ci)
        c.value = f"={gcl(ci)}{DATA_START+MAX_ROWS-1}"
        c.fill = fp(L_GOLD); c.font = fn(bold=True, sz=10)
        c.number_format = MONEY; c.alignment = al(h="right"); c.border = bdr_all

    # ── Notes row ───────────────────────────────────────────────────────
    NR = TR + 2
    ws_f.merge_cells(f"A{NR}:{gcl(40)}{NR}")
    fsp_note = (
        "CAEP costs: R57.5k onboarding (M1), tiered hosting R23k/34.5k/46k per month, 0.25% AUM fee annually."
        if mode == "caep" else
        "FINOVA CAPITAL costs: 20% of performance fees + 50% of platform fees (no fixed monthly hosting)."
    )
    ws_f.cell(row=NR, column=1,
        value=(
            "NOTES:  "
            "① Expand group D:X (click + above column D) to view per-client detail.  "
            "② Performance fee = simplified monthly accrual (AUM × monthly_return × perf_rate).  "
            "③ Exchange fee revenue = total contributions × exch_fee_rate × exch_fee_share (Variables).  "
            "   Applies to ZAR/USDT, USDT/BTC, and USDT/USDPC conversions.  "
            "④ Returns: 3-tier: 60% p.a. Y1-3 → 10% Y4 → 50% Y5-8 (next BTC cycle); all adjustable in Variables.  "
            "⑤ Reserve provision accumulates monthly (Y1-3) in the Reserve Fund Balance column; "
            "   no automated drawdown — compare Reserve Fund Bal vs Cumulative Cash to assess Y4 coverage.  "
            f"⑥ {fsp_note}"
        )).fill = fp(L_GOLD)
    ws_f.cell(row=NR, column=1).font = fn(sz=9, ital=True, col="555555")
    ws_f.cell(row=NR, column=1).alignment = al(h="left", wrap=True)
    ws_f.row_dimensions[NR].height = 75


# ═══════════════════════════════════════════════════════════════════════════
# BUILD CHARTS SHEET
# ═══════════════════════════════════════════════════════════════════════════
def build_charts(ws_ch, ws_f, mode):
    ws_ch.sheet_view.showGridLines = False
    ws_ch.merge_cells("A1:F1")
    fsp_label = "CAEP" if mode == "caep" else "FINOVA CAPITAL"
    c = ws_ch.cell(row=1, column=1,
        value=f"BitWealth (Pty) Ltd — 96-Month Visual Summary  ·  FSP: {fsp_label}")
    c.fill = fp(GOLD if mode == "caep" else ORANGE)
    c.font = Font(bold=True, sz=14, color=WHITE, name="Calibri")
    c.alignment = al(h="center"); ws_ch.row_dimensions[1].height = 24

    cats = Reference(ws_f, min_col=2,
                     min_row=DATA_START, max_row=DATA_START + MAX_ROWS - 1)

    # ① AUM growth (col 30 = AD)
    ch1 = LineChart()
    ch1.title = "Total AUM Growth (ZAR)"; ch1.style = 10
    ch1.y_axis.title = "ZAR"; ch1.x_axis.title = "Month"
    ch1.height = 14; ch1.width = 30
    ch1.add_data(Reference(ws_f, min_col=30, min_row=HDR_ROW,
                           max_row=DATA_START + MAX_ROWS - 1), titles_from_data=True)
    ch1.series[0].graphicalProperties.line.solidFill = GOLD
    ch1.series[0].graphicalProperties.line.width = 25000
    ch1.set_categories(cats)
    ws_ch.add_chart(ch1, "A3")

    # ② Revenue vs Costs (cols 29=AC and 36=AJ)
    ch2 = LineChart()
    ch2.title = "Monthly Revenue vs Costs (ZAR)"; ch2.style = 10
    ch2.y_axis.title = "ZAR"; ch2.height = 14; ch2.width = 30
    ch2.add_data(Reference(ws_f, min_col=29, min_row=HDR_ROW,
                           max_row=DATA_START + MAX_ROWS - 1), titles_from_data=True)
    ch2.add_data(Reference(ws_f, min_col=36, min_row=HDR_ROW,
                           max_row=DATA_START + MAX_ROWS - 1), titles_from_data=True)
    ch2.series[0].graphicalProperties.line.solidFill = GREEN_D
    ch2.series[0].graphicalProperties.line.width = 20000
    ch2.series[1].graphicalProperties.line.solidFill = RED_D
    ch2.series[1].graphicalProperties.line.width = 20000
    ch2.set_categories(cats)
    ws_ch.add_chart(ch2, "A32")

    # ③ Cumulative cash + Reserve fund (cols 39=AM, 40=AN)
    ch3 = LineChart()
    ch3.title = "Cumulative Cashflow & Reserve Fund (ZAR)"; ch3.style = 10
    ch3.y_axis.title = "ZAR"; ch3.height = 14; ch3.width = 30
    ch3.add_data(Reference(ws_f, min_col=39, min_row=HDR_ROW,
                           max_row=DATA_START + MAX_ROWS - 1), titles_from_data=True)
    ch3.add_data(Reference(ws_f, min_col=40, min_row=HDR_ROW,
                           max_row=DATA_START + MAX_ROWS - 1), titles_from_data=True)
    ch3.series[0].graphicalProperties.line.solidFill = TEAL
    ch3.series[0].graphicalProperties.line.width = 20000
    ch3.series[1].graphicalProperties.line.solidFill = "BF360C"
    ch3.series[1].graphicalProperties.line.width = 20000
    ch3.set_categories(cats)
    ws_ch.add_chart(ch3, "A61")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN — build both workbooks
# ═══════════════════════════════════════════════════════════════════════════
def build_workbook(mode):
    wb = Workbook()

    ws_v = wb.active
    ws_v.title = "VARIABLES"
    V = build_variables(ws_v, mode)

    ws_f = wb.create_sheet("MONTHLY FORECAST")
    build_forecast(ws_f, V, mode)

    ws_ch = wb.create_sheet("CHARTS")
    build_charts(ws_ch, ws_f, mode)

    ws_v.sheet_properties.tabColor  = GOLD if mode == "caep" else ORANGE
    ws_f.sheet_properties.tabColor  = TEAL
    ws_ch.sheet_properties.tabColor = GREEN_D

    return wb


OUTPUTS = [
    ("caep",   "BitWealth_Cashflow_Forecast_v4.0.xlsx"),
    ("finova", "BitWealth_Cashflow_Forecast_v4.0_Finova.xlsx"),
]

for mode, filename in OUTPUTS:
    wb = build_workbook(mode)
    path = OUT_DIR / filename
    wb.save(path)
    print(f"Saved ({mode:6s}): {path}")

print()
print("Key changes from v3:")
print("  · 96 months (8 years) — 2 extra bull/bear cycles")
print("  · Returns: 60% Y1-3 → 10% Y4 → 50% Y5-8 (all Variables, all adjustable)")
print("  · Exchange fee revenue: contributions × 0.1% × 50% VALR deal (new revenue column)")
print("  · Finova variant: 20% perf share + 50% plat share; no fixed FSP costs")
print("  · All v3 variable values carried across")
print()
print("Variables read from v3 and applied:")
print("  Returns:    60% Y1-3 | 10% Y4 | perf period annual (12m)")
print("  CAEP costs: R57.5k onboard | R23k/34.5k/46k hosting | 0.25% AUM fee")
print("  Client 1:   R300k lump + R200k/m | 0.75% plat | 2.5% perf")
print("  Client 2:   R40k/m | 10% perf")
print("  Client 3:   R5m lump | 10% perf")
print("  Client 4:   R20k/m | 2.5% perf")
print("  New clients: 1/m from month 3 | 48 max | R50k lump + R5k/m | 10% perf")
